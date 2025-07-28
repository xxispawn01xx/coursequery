"""
Conversation Analytics for Course Assistant
Tracks learning progress, conversation patterns, and model improvements.
"""

import json
import os
import pandas as pd
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
from collections import defaultdict, Counter
import logging

logger = logging.getLogger(__name__)

class ConversationAnalytics:
    """Analyzes conversation patterns and learning progress."""
    
    def __init__(self, conversations_dir: str = "./conversations"):
        self.conversations_dir = conversations_dir
        os.makedirs(conversations_dir, exist_ok=True)
    
    def load_all_conversations(self) -> List[Dict[str, Any]]:
        """Load all conversation data from files."""
        all_conversations = []
        
        if not os.path.exists(self.conversations_dir):
            return all_conversations
        
        for filename in os.listdir(self.conversations_dir):
            if filename.endswith('.jsonl'):
                filepath = os.path.join(self.conversations_dir, filename)
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        for line in f:
                            if line.strip():
                                conversation = json.loads(line)
                                conversation['course'] = filename.replace('_conversations.jsonl', '')
                                all_conversations.append(conversation)
                except Exception as e:
                    logger.error(f"Error loading {filename}: {e}")
        
        return all_conversations
    
    def get_analytics_summary(self) -> Dict[str, Any]:
        """Generate comprehensive analytics summary."""
        conversations = self.load_all_conversations()
        
        if not conversations:
            return {
                'total_conversations': 0,
                'courses': {},
                'learning_readiness': 0,
                'activity_patterns': {},
                'question_categories': {},
                'response_quality_trends': []
            }
        
        # Basic stats
        total_conversations = len(conversations)
        courses = self._analyze_by_course(conversations)
        
        # Learning readiness (conversations ready for fine-tuning)
        learning_readiness = sum(len(course_data['conversations']) // 10 for course_data in courses.values())
        
        # Activity patterns
        activity_patterns = self._analyze_activity_patterns(conversations)
        
        # Question categorization
        question_categories = self._categorize_questions(conversations)
        
        # Response quality trends (placeholder for future implementation)
        response_quality_trends = self._analyze_response_trends(conversations)
        
        return {
            'total_conversations': total_conversations,
            'courses': courses,
            'learning_readiness': learning_readiness,
            'activity_patterns': activity_patterns,
            'question_categories': question_categories,
            'response_quality_trends': response_quality_trends,
            'last_updated': datetime.now().isoformat()
        }
    
    def _analyze_by_course(self, conversations: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
        """Analyze conversations grouped by course."""
        course_data = defaultdict(lambda: {
            'conversations': [],
            'total_count': 0,
            'learning_triggers': 0,
            'avg_response_length': 0,
            'common_topics': [],
            'last_activity': None
        })
        
        for conv in conversations:
            course = conv.get('course', 'default')
            course_data[course]['conversations'].append(conv)
            course_data[course]['total_count'] += 1
            
            # Track last activity
            timestamp = conv.get('timestamp')
            if timestamp:
                if not course_data[course]['last_activity'] or timestamp > course_data[course]['last_activity']:
                    course_data[course]['last_activity'] = timestamp
        
        # Calculate derived metrics
        for course, data in course_data.items():
            data['learning_triggers'] = data['total_count'] // 10
            
            # Average response length
            response_lengths = [len(conv.get('answer', '')) for conv in data['conversations']]
            data['avg_response_length'] = sum(response_lengths) // len(response_lengths) if response_lengths else 0
            
            # Common topics (extract keywords from questions)
            questions = [conv.get('question', '') for conv in data['conversations']]
            data['common_topics'] = self._extract_common_topics(questions)
        
        return dict(course_data)
    
    def _analyze_activity_patterns(self, conversations: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Analyze when users are most active."""
        if not conversations:
            return {}
        
        # Parse timestamps and extract patterns
        timestamps = []
        for conv in conversations:
            try:
                timestamp = datetime.fromisoformat(conv.get('timestamp', ''))
                timestamps.append(timestamp)
            except:
                continue
        
        if not timestamps:
            return {}
        
        # Hour of day patterns
        hours = [ts.hour for ts in timestamps]
        hour_counts = Counter(hours)
        
        # Day of week patterns
        weekdays = [ts.strftime('%A') for ts in timestamps]
        weekday_counts = Counter(weekdays)
        
        # Recent activity (last 7 days)
        recent_cutoff = datetime.now() - timedelta(days=7)
        recent_activity = len([ts for ts in timestamps if ts > recent_cutoff])
        
        return {
            'peak_hours': dict(hour_counts.most_common(3)),
            'peak_days': dict(weekday_counts.most_common(3)),
            'recent_activity_7d': recent_activity,
            'total_days_active': len(set(ts.date() for ts in timestamps))
        }
    
    def _categorize_questions(self, conversations: List[Dict[str, Any]]) -> Dict[str, int]:
        """Categorize questions by type/topic."""
        categories = {
            'excel_spreadsheet': 0,
            'financial_analysis': 0,
            'course_content': 0,
            'calculations': 0,
            'definitions': 0,
            'examples': 0,
            'how_to': 0,
            'comparison': 0
        }
        
        keywords = {
            'excel_spreadsheet': ['excel', 'spreadsheet', 'table', 'csv'],
            'financial_analysis': ['financial', 'analysis', 'roi', 'cash flow', 'investment'],
            'course_content': ['explain', 'what is', 'define', 'meaning'],
            'calculations': ['calculate', 'formula', 'compute', 'math'],
            'definitions': ['definition', 'define', 'what is', 'meaning'],
            'examples': ['example', 'sample', 'instance', 'case study'],
            'how_to': ['how to', 'how do', 'steps', 'process'],
            'comparison': ['compare', 'difference', 'versus', 'vs', 'better']
        }
        
        for conv in conversations:
            question = conv.get('question', '').lower()
            for category, words in keywords.items():
                if any(word in question for word in words):
                    categories[category] += 1
        
        return categories
    
    def _analyze_response_trends(self, conversations: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Analyze response quality trends over time."""
        # Placeholder for future implementation
        # Could include metrics like response length, complexity, user satisfaction
        
        if not conversations:
            return []
        
        # Group by date and calculate basic metrics
        daily_stats = defaultdict(lambda: {'count': 0, 'avg_length': 0, 'responses': []})
        
        for conv in conversations:
            try:
                timestamp = datetime.fromisoformat(conv.get('timestamp', ''))
                date_key = timestamp.strftime('%Y-%m-%d')
                daily_stats[date_key]['count'] += 1
                daily_stats[date_key]['responses'].append(len(conv.get('answer', '')))
            except:
                continue
        
        # Calculate averages
        trends = []
        for date, stats in sorted(daily_stats.items()):
            if stats['responses']:
                avg_length = sum(stats['responses']) // len(stats['responses'])
                trends.append({
                    'date': date,
                    'conversation_count': stats['count'],
                    'avg_response_length': avg_length
                })
        
        return trends[-30:]  # Last 30 days
    
    def _extract_common_topics(self, questions: List[str]) -> List[str]:
        """Extract common topics from questions."""
        # Simple keyword extraction
        stop_words = {'the', 'is', 'at', 'which', 'on', 'a', 'an', 'and', 'or', 'but', 'in', 'with', 'to', 'for', 'of', 'as', 'by'}
        
        all_words = []
        for question in questions:
            words = question.lower().split()
            filtered_words = [word.strip('.,?!') for word in words if len(word) > 3 and word.lower() not in stop_words]
            all_words.extend(filtered_words)
        
        word_counts = Counter(all_words)
        return [word for word, count in word_counts.most_common(5) if count > 1]
    
    def export_analytics_excel(self) -> bytes:
        """Export analytics to Excel format."""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        analytics = self.get_analytics_summary()
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Headers
        ws_summary['A1'] = "Course Learning Analytics"
        ws_summary['A1'].font = Font(bold=True, size=16)
        
        ws_summary['A3'] = "Total Conversations"
        ws_summary['B3'] = analytics['total_conversations']
        
        ws_summary['A4'] = "Learning Triggers Ready"
        ws_summary['B4'] = analytics['learning_readiness']
        
        ws_summary['A5'] = "Active Courses"
        ws_summary['B5'] = len(analytics['courses'])
        
        # Course details sheet
        if analytics['courses']:
            ws_courses = wb.create_sheet(title="Courses")
            ws_courses['A1'] = "Course"
            ws_courses['B1'] = "Conversations"
            ws_courses['C1'] = "Learning Triggers"
            ws_courses['D1'] = "Avg Response Length"
            ws_courses['E1'] = "Last Activity"
            
            for i, (course, data) in enumerate(analytics['courses'].items(), start=2):
                ws_courses[f'A{i}'] = course
                ws_courses[f'B{i}'] = data['total_count']
                ws_courses[f'C{i}'] = data['learning_triggers']
                ws_courses[f'D{i}'] = data['avg_response_length']
                ws_courses[f'E{i}'] = data['last_activity']
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()