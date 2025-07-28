"""
Model Evaluation Metrics for ML Pipeline Validation
Tracks performance, quality, and resource utilization metrics that data scientists use.
"""

import time
import psutil
import json
import os
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import logging
from collections import deque, defaultdict
import threading

logger = logging.getLogger(__name__)

class ModelEvaluationMetrics:
    """Comprehensive model evaluation and monitoring system."""
    
    def __init__(self, metrics_dir: str = "./metrics"):
        self.metrics_dir = metrics_dir
        os.makedirs(metrics_dir, exist_ok=True)
        
        # Performance tracking
        self.response_times = deque(maxlen=1000)  # Last 1000 responses
        self.token_counts = deque(maxlen=1000)
        self.memory_usage = deque(maxlen=100)
        
        # Quality metrics
        self.response_lengths = deque(maxlen=1000)
        self.conversation_scores = deque(maxlen=100)
        
        # Resource monitoring
        self.gpu_usage = deque(maxlen=100)
        self.cpu_usage = deque(maxlen=100)
        
        # Model-specific metrics
        self.model_metrics = defaultdict(lambda: {
            'total_queries': 0,
            'avg_response_time': 0,
            'avg_tokens_per_second': 0,
            'error_rate': 0,
            'last_used': None
        })
        
        # Start background monitoring
        self._start_resource_monitoring()
    
    def start_query_timing(self) -> float:
        """Start timing a query. Returns start timestamp."""
        return time.time()
    
    def end_query_timing(self, start_time: float, model_name: str, 
                        input_text: str, output_text: str, 
                        error: Optional[str] = None) -> Dict[str, Any]:
        """End timing and record comprehensive metrics."""
        end_time = time.time()
        response_time = end_time - start_time
        
        # Calculate token metrics
        input_tokens = len(input_text.split())  # Approximate token count
        output_tokens = len(output_text.split())
        total_tokens = input_tokens + output_tokens
        tokens_per_second = total_tokens / response_time if response_time > 0 else 0
        
        # Record performance metrics
        self.response_times.append(response_time)
        self.token_counts.append(total_tokens)
        self.response_lengths.append(len(output_text))
        
        # Update model-specific metrics
        self.model_metrics[model_name]['total_queries'] += 1
        self.model_metrics[model_name]['last_used'] = datetime.now().isoformat()
        
        # Calculate running averages
        if error:
            self.model_metrics[model_name]['error_rate'] = (
                self.model_metrics[model_name].get('error_rate', 0) * 0.9 + 0.1
            )
        else:
            self.model_metrics[model_name]['error_rate'] = (
                self.model_metrics[model_name].get('error_rate', 0) * 0.9
            )
        
        # Update averages
        current_avg_time = self.model_metrics[model_name]['avg_response_time']
        self.model_metrics[model_name]['avg_response_time'] = (
            current_avg_time * 0.9 + response_time * 0.1
        )
        
        current_avg_tps = self.model_metrics[model_name]['avg_tokens_per_second']
        self.model_metrics[model_name]['avg_tokens_per_second'] = (
            current_avg_tps * 0.9 + tokens_per_second * 0.1
        )
        
        # Calculate quality metrics
        quality_score = self._calculate_response_quality(input_text, output_text)
        
        metrics = {
            'timestamp': datetime.now().isoformat(),
            'model_name': model_name,
            'response_time_ms': response_time * 1000,
            'input_tokens': input_tokens,
            'output_tokens': output_tokens,
            'total_tokens': total_tokens,
            'tokens_per_second': tokens_per_second,
            'response_length_chars': len(output_text),
            'quality_score': quality_score,
            'error': error,
            'memory_usage_mb': self._get_current_memory_usage(),
            'cpu_usage_percent': self._get_current_cpu_usage()
        }
        
        # Save metrics
        self._save_metrics(metrics)
        
        return metrics
    
    def _calculate_response_quality(self, input_text: str, output_text: str) -> float:
        """Calculate response quality score (0-1)."""
        if not output_text or not input_text:
            return 0.0
        
        quality_score = 0.0
        
        # Length appropriateness (not too short, not too long)
        response_ratio = len(output_text) / len(input_text)
        if 0.5 <= response_ratio <= 10:  # Reasonable response length
            quality_score += 0.3
        
        # Vocabulary diversity
        output_words = set(output_text.lower().split())
        if len(output_words) > 10:  # Diverse vocabulary
            quality_score += 0.2
        
        # Coherence (simple check for sentence structure)
        sentences = output_text.split('.')
        if len(sentences) >= 2:  # Multiple sentences suggest structure
            quality_score += 0.2
        
        # Relevance (keyword overlap)
        input_words = set(input_text.lower().split())
        common_words = input_words.intersection(output_words)
        if len(common_words) > 0:
            relevance = min(len(common_words) / len(input_words), 0.3)
            quality_score += relevance
        
        return min(quality_score, 1.0)
    
    def _get_current_memory_usage(self) -> float:
        """Get current memory usage in MB."""
        try:
            process = psutil.Process()
            return process.memory_info().rss / 1024 / 1024  # MB
        except:
            return 0.0
    
    def _get_current_cpu_usage(self) -> float:
        """Get current CPU usage percentage."""
        try:
            return psutil.cpu_percent(interval=0.1)
        except:
            return 0.0
    
    def _get_gpu_usage(self) -> Dict[str, float]:
        """Get GPU usage if available."""
        try:
            import torch
            if torch.cuda.is_available():
                gpu_memory = torch.cuda.memory_allocated() / 1024**3  # GB
                gpu_util = torch.cuda.utilization() if hasattr(torch.cuda, 'utilization') else 0
                return {
                    'gpu_memory_gb': gpu_memory,
                    'gpu_utilization_percent': gpu_util
                }
        except:
            pass
        return {'gpu_memory_gb': 0, 'gpu_utilization_percent': 0}
    
    def _start_resource_monitoring(self):
        """Start background thread to monitor system resources."""
        def monitor():
            while True:
                try:
                    # CPU and memory
                    self.cpu_usage.append(self._get_current_cpu_usage())
                    self.memory_usage.append(self._get_current_memory_usage())
                    
                    # GPU if available
                    gpu_metrics = self._get_gpu_usage()
                    self.gpu_usage.append(gpu_metrics.get('gpu_utilization_percent', 0))
                    
                    time.sleep(5)  # Monitor every 5 seconds
                except Exception as e:
                    logger.error(f"Resource monitoring error: {e}")
                    time.sleep(10)
        
        monitor_thread = threading.Thread(target=monitor, daemon=True)
        monitor_thread.start()
    
    def _save_metrics(self, metrics: Dict[str, Any]):
        """Save metrics to file."""
        metrics_file = os.path.join(self.metrics_dir, f"metrics_{datetime.now().strftime('%Y%m%d')}.jsonl")
        try:
            with open(metrics_file, 'a', encoding='utf-8') as f:
                f.write(json.dumps(metrics) + '\n')
        except Exception as e:
            logger.error(f"Error saving metrics: {e}")
    
    def get_performance_summary(self) -> Dict[str, Any]:
        """Get comprehensive performance summary."""
        if not self.response_times:
            return {"status": "no_data", "message": "No performance data available yet"}
        
        # Calculate statistics
        response_times_list = list(self.response_times)
        token_counts_list = list(self.token_counts)
        response_lengths_list = list(self.response_lengths)
        
        summary = {
            'timestamp': datetime.now().isoformat(),
            'performance_metrics': {
                'avg_response_time_ms': sum(response_times_list) / len(response_times_list) * 1000,
                'p95_response_time_ms': sorted(response_times_list)[int(0.95 * len(response_times_list))] * 1000,
                'p99_response_time_ms': sorted(response_times_list)[int(0.99 * len(response_times_list))] * 1000,
                'min_response_time_ms': min(response_times_list) * 1000,
                'max_response_time_ms': max(response_times_list) * 1000,
                'total_queries': len(response_times_list)
            },
            'throughput_metrics': {
                'avg_tokens_per_query': sum(token_counts_list) / len(token_counts_list),
                'avg_tokens_per_second': sum(token_counts_list) / sum(response_times_list),
                'avg_response_length_chars': sum(response_lengths_list) / len(response_lengths_list)
            },
            'resource_metrics': {
                'avg_memory_usage_mb': sum(self.memory_usage) / len(self.memory_usage) if self.memory_usage else 0,
                'avg_cpu_usage_percent': sum(self.cpu_usage) / len(self.cpu_usage) if self.cpu_usage else 0,
                'avg_gpu_usage_percent': sum(self.gpu_usage) / len(self.gpu_usage) if self.gpu_usage else 0,
                'current_memory_mb': self._get_current_memory_usage(),
                'current_cpu_percent': self._get_current_cpu_usage()
            },
            'quality_metrics': {
                'avg_quality_score': sum(self.conversation_scores) / len(self.conversation_scores) if self.conversation_scores else 0,
                'response_consistency': self._calculate_response_consistency()
            },
            'model_breakdown': dict(self.model_metrics)
        }
        
        # Add GPU metrics if available
        gpu_metrics = self._get_gpu_usage()
        summary['resource_metrics'].update(gpu_metrics)
        
        return summary
    
    def _calculate_response_consistency(self) -> float:
        """Calculate how consistent response lengths are."""
        if len(self.response_lengths) < 2:
            return 1.0
        
        lengths = list(self.response_lengths)
        avg_length = sum(lengths) / len(lengths)
        variance = sum((l - avg_length) ** 2 for l in lengths) / len(lengths)
        std_dev = variance ** 0.5
        
        # Normalize by average (coefficient of variation)
        consistency = 1.0 - min(std_dev / avg_length, 1.0) if avg_length > 0 else 0.0
        return consistency
    
    def get_model_comparison(self) -> Dict[str, Any]:
        """Compare performance across different models."""
        if not self.model_metrics:
            return {"status": "no_data", "message": "No model data available"}
        
        comparison = {}
        for model_name, metrics in self.model_metrics.items():
            comparison[model_name] = {
                'queries_handled': metrics['total_queries'],
                'avg_response_time_ms': metrics['avg_response_time'] * 1000,
                'avg_tokens_per_second': metrics['avg_tokens_per_second'],
                'error_rate_percent': metrics['error_rate'] * 100,
                'last_used': metrics['last_used'],
                'performance_grade': self._calculate_performance_grade(metrics)
            }
        
        return {
            'model_comparison': comparison,
            'best_performer': self._find_best_performer(comparison),
            'recommendations': self._generate_performance_recommendations(comparison)
        }
    
    def _calculate_performance_grade(self, metrics: Dict[str, Any]) -> str:
        """Calculate overall performance grade for a model."""
        score = 0
        
        # Response time (lower is better)
        if metrics['avg_response_time'] < 1.0:
            score += 25
        elif metrics['avg_response_time'] < 3.0:
            score += 20
        elif metrics['avg_response_time'] < 5.0:
            score += 15
        
        # Tokens per second (higher is better)
        if metrics['avg_tokens_per_second'] > 50:
            score += 25
        elif metrics['avg_tokens_per_second'] > 20:
            score += 20
        elif metrics['avg_tokens_per_second'] > 10:
            score += 15
        
        # Error rate (lower is better)
        if metrics['error_rate'] < 0.01:
            score += 25
        elif metrics['error_rate'] < 0.05:
            score += 20
        elif metrics['error_rate'] < 0.1:
            score += 15
        
        # Usage (more queries = more confidence in metrics)
        if metrics['total_queries'] > 100:
            score += 25
        elif metrics['total_queries'] > 50:
            score += 20
        elif metrics['total_queries'] > 10:
            score += 15
        
        if score >= 90:
            return "A"
        elif score >= 80:
            return "B"
        elif score >= 70:
            return "C"
        elif score >= 60:
            return "D"
        else:
            return "F"
    
    def _find_best_performer(self, comparison: Dict[str, Any]) -> Optional[str]:
        """Find the best performing model."""
        if not comparison:
            return None
        
        best_model = None
        best_score = -1
        
        for model_name, metrics in comparison.items():
            # Simple scoring: prioritize speed and low error rate
            speed_score = min(metrics['avg_tokens_per_second'] / 10, 10)  # Normalize to 0-10
            reliability_score = max(0, 10 - metrics['error_rate_percent'])  # 0-10
            total_score = speed_score + reliability_score
            
            if total_score > best_score:
                best_score = total_score
                best_model = model_name
        
        return best_model
    
    def _generate_performance_recommendations(self, comparison: Dict[str, Any]) -> List[str]:
        """Generate performance improvement recommendations."""
        recommendations = []
        
        if not comparison:
            return ["No data available for recommendations"]
        
        # Analyze patterns
        avg_response_time = sum(m['avg_response_time_ms'] for m in comparison.values()) / len(comparison)
        avg_error_rate = sum(m['error_rate_percent'] for m in comparison.values()) / len(comparison)
        
        if avg_response_time > 3000:
            recommendations.append("Consider model optimization - response times are high")
        
        if avg_error_rate > 5:
            recommendations.append("High error rates detected - check model authentication and resources")
        
        # Check for underutilized models
        low_usage_models = [name for name, metrics in comparison.items() if metrics['queries_handled'] < 10]
        if low_usage_models:
            recommendations.append(f"Models with low usage: {', '.join(low_usage_models)} - consider testing more")
        
        # Resource recommendations
        if hasattr(self, 'memory_usage') and self.memory_usage:
            avg_memory = sum(self.memory_usage) / len(self.memory_usage)
            if avg_memory > 8000:  # 8GB
                recommendations.append("High memory usage detected - consider model quantization")
        
        if not recommendations:
            recommendations.append("Performance looks good - continue monitoring")
        
        return recommendations
    
    def export_metrics_report(self) -> bytes:
        """Export comprehensive metrics report as Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            import io
            
            wb = Workbook()
            
            # Performance summary sheet
            ws_summary = wb.active
            ws_summary.title = "Performance Summary"
            
            summary = self.get_performance_summary()
            
            # Headers
            ws_summary['A1'] = "Model Performance Report"
            ws_summary['A1'].font = Font(bold=True, size=16)
            
            row = 3
            for category, metrics in summary.items():
                if isinstance(metrics, dict):
                    ws_summary[f'A{row}'] = category.replace('_', ' ').title()
                    ws_summary[f'A{row}'].font = Font(bold=True)
                    row += 1
                    
                    for metric, value in metrics.items():
                        ws_summary[f'B{row}'] = metric.replace('_', ' ').title()
                        ws_summary[f'C{row}'] = value
                        row += 1
                    row += 1
            
            # Model comparison sheet
            comparison = self.get_model_comparison()
            if comparison.get('model_comparison'):
                ws_models = wb.create_sheet(title="Model Comparison")
                ws_models['A1'] = "Model"
                ws_models['B1'] = "Queries"
                ws_models['C1'] = "Avg Response Time (ms)"
                ws_models['D1'] = "Tokens/Second"
                ws_models['E1'] = "Error Rate (%)"
                ws_models['F1'] = "Grade"
                
                for i, (model, metrics) in enumerate(comparison['model_comparison'].items(), start=2):
                    ws_models[f'A{i}'] = model
                    ws_models[f'B{i}'] = metrics['queries_handled']
                    ws_models[f'C{i}'] = round(metrics['avg_response_time_ms'], 2)
                    ws_models[f'D{i}'] = round(metrics['avg_tokens_per_second'], 2)
                    ws_models[f'E{i}'] = round(metrics['error_rate_percent'], 2)
                    ws_models[f'F{i}'] = metrics['performance_grade']
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting metrics report: {e}")
            return b""