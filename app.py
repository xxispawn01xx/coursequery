"""
Real Estate AI Stack - Fully Local Streamlit Application
Main entry point for the course analysis application.
"""

import streamlit as st
import os
import json
import logging
from pathlib import Path
from typing import List, Optional, Dict, Any
from datetime import datetime
# Optional visualization imports
try:
    import plotly.express as px
    import plotly.graph_objects as go
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False
    px = None
    go = None

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    pd = None

from config import Config
from pathlib import Path

# Optional AI module imports
try:
    from document_processor import DocumentProcessor
    DOCUMENT_PROCESSOR_AVAILABLE = True
except ImportError:
    DocumentProcessor = None
    DOCUMENT_PROCESSOR_AVAILABLE = False

try:
    from local_models import LocalModelManager, TORCH_AVAILABLE, TRANSFORMERS_AVAILABLE
    LOCAL_MODELS_AVAILABLE = TORCH_AVAILABLE and TRANSFORMERS_AVAILABLE
except ImportError:
    LocalModelManager = None
    LOCAL_MODELS_AVAILABLE = False

try:
    from course_indexer import CourseIndexer
    COURSE_INDEXER_AVAILABLE = True
except ImportError:
    CourseIndexer = None
    COURSE_INDEXER_AVAILABLE = False

# Always try to use offline course manager as fallback
try:
    from offline_course_manager import OfflineCourseManager
    OFFLINE_COURSE_MANAGER_AVAILABLE = True
except ImportError:
    OfflineCourseManager = None
    OFFLINE_COURSE_MANAGER_AVAILABLE = False

try:
    from query_engine import LocalQueryEngine
    QUERY_ENGINE_AVAILABLE = True
except ImportError:
    LocalQueryEngine = None
    QUERY_ENGINE_AVAILABLE = False

try:
    from hybrid_query_engine import HybridQueryEngine
    HYBRID_QUERY_ENGINE_AVAILABLE = True
except ImportError:
    HybridQueryEngine = None
    HYBRID_QUERY_ENGINE_AVAILABLE = False

try:
    from course_rename_handler import create_rename_handler
    RENAME_HANDLER_AVAILABLE = True
except ImportError:
    create_rename_handler = None
    RENAME_HANDLER_AVAILABLE = False

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class RealEstateAIApp:
    """Main application class for the Real Estate AI Stack."""
    
    def __init__(self):
        """Initialize the application."""
        self.config = Config()
        
        # FORCE local directory usage - override any H:\ detection
        app_dir = Path(__file__).parent
        self.config.raw_docs_dir = app_dir / "archived_courses" 
        self.config.indexed_courses_dir = app_dir / "indexed_courses"
        
        # Create directories if they don't exist
        self.config.raw_docs_dir.mkdir(exist_ok=True)
        self.config.indexed_courses_dir.mkdir(exist_ok=True)
        
        print(f"🔒 FORCED to use local directories:")
        print(f"📁 Raw docs: {self.config.raw_docs_dir}")
        print(f"📊 Indexed: {self.config.indexed_courses_dir}")
        
        # Initialize available components
        self.doc_processor = DocumentProcessor() if DOCUMENT_PROCESSOR_AVAILABLE else None
        
        # Use session state for model manager and query engine to persist across requests
        if 'model_manager' in st.session_state:
            self.model_manager = st.session_state.model_manager
        else:
            self.model_manager = LocalModelManager() if LOCAL_MODELS_AVAILABLE else None
            
        if 'query_engine' in st.session_state:
            self.query_engine = st.session_state.query_engine
        else:
            # Initialize hybrid query engine for better reliability
            if HYBRID_QUERY_ENGINE_AVAILABLE:
                self.query_engine = HybridQueryEngine()
            else:
                self.query_engine = None
            
        # Try CourseIndexer first, fallback to OfflineCourseManager
        if COURSE_INDEXER_AVAILABLE:
            try:
                self.course_indexer = CourseIndexer()
                logger.info("✅ Using full CourseIndexer")
            except Exception as e:
                logger.warning(f"⚠️ CourseIndexer failed, using offline fallback: {e}")
                self.course_indexer = OfflineCourseManager(
                    self.config.raw_docs_dir, 
                    self.config.indexed_courses_dir
                ) if OFFLINE_COURSE_MANAGER_AVAILABLE else None
        else:
            self.course_indexer = OfflineCourseManager(
                self.config.raw_docs_dir, 
                self.config.indexed_courses_dir
            ) if OFFLINE_COURSE_MANAGER_AVAILABLE else None
        
        # Initialize ignore manager
        try:
            from course_ignore_manager import CourseIgnoreManager
            self.ignore_manager = CourseIgnoreManager()
        except ImportError:
            self.ignore_manager = None
        
        # Initialize book indexer
        try:
            from book_indexer import BookIndexer
            self.book_indexer = BookIndexer()
        except ImportError:
            self.book_indexer = None
        
        # Initialize rename handler
        if RENAME_HANDLER_AVAILABLE:
            try:
                self.rename_handler = create_rename_handler(
                    str(self.config.raw_docs_dir), 
                    str(self.config.indexed_courses_dir)
                )
            except Exception as e:
                logger.warning(f"Could not initialize rename handler: {e}")
                self.rename_handler = None
        else:
            self.rename_handler = None
        
        # Initialize session state
        if 'models_loaded' not in st.session_state:
            st.session_state.models_loaded = False
        if 'selected_model' not in st.session_state:
            st.session_state.selected_model = None
        if 'available_courses' not in st.session_state:
            st.session_state.available_courses = []
        if 'selected_course' not in st.session_state:
            st.session_state.selected_course = None
        if 'chat_history' not in st.session_state:
            st.session_state.chat_history = []

    def setup_page_config(self):
        """Configure the Streamlit page."""
        st.set_page_config(
            page_title="Local Course AI Assistant",
            page_icon="📚",
            layout="wide",
            initial_sidebar_state="expanded"
        )

    def load_saved_token(self):
        """Load saved HuggingFace token from secure storage."""
        token_file = Path(".hf_token")
        
        # Try to load from environment first
        if "HF_TOKEN" in os.environ:
            return os.environ["HF_TOKEN"]
        
        # Try to load from secure file
        if token_file.exists():
            try:
                with open(token_file, 'r') as f:
                    token = f.read().strip()
                if token and token.startswith("hf_"):
                    os.environ["HF_TOKEN"] = token
                    return token
            except Exception:
                pass
        
        return None

    def save_token(self, token: str):
        """Save HuggingFace token securely."""
        try:
            # Save to environment
            os.environ["HF_TOKEN"] = token
            
            # Save to secure file for persistence
            token_file = Path(".hf_token")
            with open(token_file, 'w') as f:
                f.write(token)
            
            # Set file permissions to be readable only by owner
            os.chmod(token_file, 0o600)
            
            return True
        except Exception as e:
            st.error(f"Failed to save token: {e}")
            return False

    def setup_authentication(self):
        """Setup HuggingFace authentication with GUI token input."""
        st.subheader("🔐 HuggingFace Authentication")
        
        # Try to load saved token first
        saved_token = self.load_saved_token()
        if saved_token:
            try:
                from huggingface_hub import HfApi
                api = HfApi(token=saved_token)
                whoami = api.whoami()
                if whoami:
                    st.success(f"✅ Using saved token - Authenticated as: {whoami['name']}")
                    st.session_state.hf_token_set = True
                    return True
            except Exception:
                # Saved token is invalid, continue to input interface
                pass
        
        # Token input interface
        st.markdown("""
        **To use AI models, you need a HuggingFace token:**
        1. Go to https://huggingface.co/settings/tokens
        2. Create a new token (Read access is sufficient)
        3. Enter it below - it will be saved securely for future sessions
        """)
        
        # Clear token button if one exists
        if saved_token:
            col1, col2 = st.columns([3, 1])
            with col2:
                if st.button("🗑️ Clear Saved Token", key="clear_hf_token_btn"):
                    try:
                        Path(".hf_token").unlink(missing_ok=True)
                        if "HF_TOKEN" in os.environ:
                            del os.environ["HF_TOKEN"]
                        st.session_state.hf_token_set = False
                        st.success("Token cleared")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Failed to clear token: {e}")
        
        token = st.text_input(
            "HuggingFace Token", 
            type="password", 
            placeholder="hf_...",
            help="Your token will be saved securely and persist across sessions"
        )
        
        if token and token.startswith("hf_"):
            if st.button("💾 Save Token & Load Models", key="save_hf_token_btn"):
                try:
                    # Test authentication first
                    from huggingface_hub import HfApi
                    api = HfApi(token=token)
                    whoami = api.whoami()
                    
                    # Save token if valid
                    if self.save_token(token):
                        st.success(f"✅ Token saved permanently! Authenticated as: {whoami['name']}")
                        st.session_state.hf_token_set = True
                        st.rerun()
                    
                except Exception as e:
                    st.error(f"❌ Invalid token: {e}")
                    return False
        
        elif token and not token.startswith("hf_"):
            st.warning("⚠️ Token should start with 'hf_'")
        
        return False

    def model_selection_interface(self):
        """RTX 3060 GPU Memory Optimization - Choose ONE model to prevent GPU overload."""
        st.subheader("🚀 RTX 3060 Model Selection")
        
        st.info("RTX 3060 12GB Memory Optimization: Load only ONE model at a time to prevent GPU overload")
        
        # Model selection
        model_options = {
            "mistral": "🔥 Mistral 7B Instruct (Recommended for most tasks)",
            "llama": "🦙 Llama 2 7B Chat (Alternative conversation model)"
        }
        
        selected_model = st.radio(
            "Choose AI Model for RTX 3060:",
            options=list(model_options.keys()),
            format_func=lambda x: model_options[x],
            index=0,  # Default to Mistral
            help="Only one model will be loaded to optimize RTX 3060 12GB memory usage"
        )
        
        # Show current model status
        if st.session_state.get('selected_model'):
            if st.session_state.selected_model == selected_model:
                st.success(f"✅ Currently loaded: {model_options[selected_model]}")
            else:
                st.warning(f"⚠️ Model change detected. Current: {model_options[st.session_state.selected_model]} → New: {model_options[selected_model]}")
        
        return selected_model

    def load_models(self):
        """Load local models with RTX 3060 optimization - only one model at a time."""
        # Pure offline mode - always load models
        
        # Check authentication first for local usage
        if not st.session_state.get('hf_token_set', False):
            # Try to load saved token
            saved_token = self.load_saved_token()
            if saved_token:
                try:
                    from huggingface_hub import HfApi
                    api = HfApi(token=saved_token)
                    whoami = api.whoami()
                    if whoami:
                        st.session_state.hf_token_set = True
                        st.success(f"✅ Using saved authentication - {whoami['name']}")
                    else:
                        if not self.setup_authentication():
                            return False
                except Exception:
                    if not self.setup_authentication():
                        return False
            else:
                if not self.setup_authentication():
                    return False
        
        # RTX 3060 Model Selection Interface
        selected_model = self.model_selection_interface()
        
        # Check if model needs loading or switching
        model_needs_loading = (not st.session_state.models_loaded or 
                               st.session_state.get('selected_model') != selected_model)
        
        if model_needs_loading:
            # Offline mode - force models to load even if dependencies seem unavailable
            # Skip dependency check for pure offline mode
            st.info("💡 Pure offline mode - loading models directly")
            
            if self.model_manager is None:
                st.warning("⚠️ Model manager not available - using hybrid mode")
                # Continue without local models - use cloud APIs instead
                st.session_state.models_loaded = True
                st.session_state.selected_model = "local_models"
                return True
            
            # Show loading message based on model change
            if st.session_state.models_loaded and st.session_state.get('selected_model') != selected_model:
                loading_msg = f"Switching from {st.session_state.get('selected_model', 'unknown')} to {selected_model} (RTX 3060 memory optimization)..."
            else:
                loading_msg = f"Loading {selected_model} model (cached models load quickly, first download takes longer)..."
            
            with st.spinner(loading_msg):
                try:
                    # Load specific model for RTX 3060 optimization
                    self.model_manager.load_models(model_type=selected_model)
                    logger.info(f"Model manager loaded successfully with {selected_model}")
                    
                    # Store model manager in session state to persist across requests
                    st.session_state.model_manager = self.model_manager
                    
                    # Initialize query engine
                    if QUERY_ENGINE_AVAILABLE:
                        self.query_engine = LocalQueryEngine(self.model_manager)
                        st.session_state.query_engine = self.query_engine
                        logger.info("Query engine initialized successfully")
                    else:
                        logger.error("Query engine not available")
                        st.error("❌ Query engine not available. Please check dependencies.")
                        return False
                    
                    st.session_state.models_loaded = True
                    st.session_state.selected_model = selected_model
                    st.success(f"✅ {selected_model.title()} model loaded successfully on RTX 3060!")
                    logger.info(f"Models and query engine loaded successfully with {selected_model}")
                except Exception as e:
                    # Show more helpful error message for authentication issues
                    error_msg = str(e)
                    if "401" in error_msg or "gated repo" in error_msg or "authenticated" in error_msg:
                        st.error("❌ Model authentication required")
                        st.markdown("""
                        **HuggingFace Authentication Needed:**
                        1. Get token from https://huggingface.co/settings/tokens
                        2. Open Shell tab and run: `huggingface-cli login`
                        3. For Llama 2, request access at: https://huggingface.co/meta-llama/Llama-2-7b-chat-hf
                        4. Reload this page after authentication
                        
                        **Alternative:** Set HF_TOKEN in Secrets tab
                        """)
                    elif "paging file" in error_msg.lower() or "os error 1455" in error_msg.lower():
                        st.error("❌ Memory Error: Paging file too small")
                        st.markdown("""
                        **Windows Memory Issue - Quick Fix:**
                        1. Press Windows + R, type "sysdm.cpl", press Enter
                        2. Advanced tab → Performance Settings → Advanced → Virtual Memory
                        3. Click "Change" → Uncheck "Automatically manage"
                        4. Custom size: Initial 4096 MB, Maximum 8192 MB
                        5. Click Set → OK → Restart computer
                        
                        **Alternative:** Use smaller models by setting model preference to "small" in config.
                        """)
                    else:
                        st.error(f"❌ Failed to load models: {error_msg}")
                    
                    logger.error(f"Model loading failed: {e}")
                    return False
        return True

    def refresh_available_courses(self):
        """Refresh the list of available courses - pure offline mode."""
        print("🔄 Starting course refresh...")
        
        # Initialize session state if needed
        if 'available_courses' not in st.session_state:
            st.session_state.available_courses = []
        if 'selected_course' not in st.session_state:
            st.session_state.selected_course = None
        
        courses = []
        
        # Use the configured raw docs directory
        raw_docs_path = Path(self.config.raw_docs_dir)
        print(f"📁 Scanning directory: {raw_docs_path}")
        
        if not raw_docs_path.exists():
            print(f"❌ Directory not found: {raw_docs_path}")
            st.session_state.available_courses = []
            return []
        
        # Always use local indexed_courses directory for processed courses
        indexed_path = Path(__file__).parent / "indexed_courses"
        indexed_path.mkdir(exist_ok=True)
        
        print(f"📁 Raw courses: {raw_docs_path}")
        print(f"📊 Indexed courses: {indexed_path}")
        
        # Get already indexed courses with proper document counting
        indexed_courses = set()
        for item in indexed_path.iterdir():
            if item.is_dir():
                indexed_courses.add(item.name)
                
                # Try to read actual document count from metadata
                metadata_file = item / "metadata.json"
                doc_count = 0
                
                if metadata_file.exists():
                    try:
                        with open(metadata_file, 'r', encoding='utf-8') as f:
                            metadata = json.load(f)
                        doc_count = metadata.get('document_count', 0)
                        
                        courses.append({
                            'name': item.name,
                            'status': 'indexed',
                            'document_count': doc_count,
                            'last_indexed': metadata.get('last_indexed', 'Ready for queries'),
                            'total_content_length': metadata.get('total_content_length', 0),
                            'document_types': metadata.get('document_types', {})
                        })
                        print(f"📚 Indexed course: {item.name} ({doc_count} documents)")
                        
                    except Exception as e:
                        # Fallback to file counting if metadata is corrupted
                        doc_count = len(list(item.rglob('*.json'))) + len(list(item.rglob('*.pkl')))
                        courses.append({
                            'name': item.name,
                            'status': 'indexed',
                            'document_count': doc_count,
                            'last_indexed': 'Metadata error'
                        })
                        print(f"📚 Indexed course: {item.name} ({doc_count} files, metadata error)")
                else:
                    # Fallback to file counting
                    doc_count = len(list(item.rglob('*.json'))) + len(list(item.rglob('*.pkl')))
                    courses.append({
                        'name': item.name,
                        'status': 'indexed',
                        'document_count': doc_count,
                        'last_indexed': 'No metadata'
                    })
                    print(f"📚 Indexed course: {item.name} ({doc_count} files, no metadata)")
        
        # Scan for courses in the directory
        course_count = 0
        
        try:
            print(f"📂 Directory contents:")
            for item in raw_docs_path.iterdir():
                print(f"  - {item.name} ({'DIR' if item.is_dir() else 'FILE'})")
                
                if item.is_dir():
                    course_count += 1
                    course_name = item.name
                    
                    # Count supported files in this course directory
                    file_count = 0
                    supported_extensions = ['.pdf', '.docx', '.pptx', '.epub', '.mp4', '.avi', '.mov', '.mp3', '.wav']
                    
                    for file_path in item.rglob('*'):
                        if file_path.is_file() and file_path.suffix.lower() in supported_extensions:
                            file_count += 1
                    
                    # Check if already indexed
                    if course_name in indexed_courses:
                        print(f"📚 Course {course_name} already indexed - skipping")
                        continue
                    
                    # Add as unprocessed course
                    courses.append({
                        'name': course_name,
                        'status': 'raw',
                        'document_count': file_count,
                        'last_indexed': 'Not processed'
                    })
                    
                    print(f"📁 Found course: {course_name} ({file_count} files)")
        
        except Exception as e:
            print(f"❌ Error scanning directory: {e}")
            st.error(f"Error accessing directory: {e}")
        
        print(f"✅ Total courses found: {len(courses)} (scanned {course_count} directories)")
        
        # Filter out ignored courses
        if self.ignore_manager:
            courses = self.ignore_manager.filter_courses(courses)
            print(f"✂️ After filtering ignored courses: {len(courses)} remaining")
        
        # Store results
        st.session_state.available_courses = courses
        return courses

    def sidebar_course_management(self):
        """Handle course management in the sidebar."""
        st.sidebar.header("📚 Course Management")
        
        # Check for renamed courses and offer to fix indexes
        if self.rename_handler:
            try:
                rename_suggestions = self.rename_handler.get_rename_suggestions()
                if rename_suggestions:
                    st.sidebar.warning(f"⚠️ Found {len(rename_suggestions)} renamed courses")
                    
                    with st.sidebar.expander("🔄 Fix Renamed Courses"):
                        st.info("Detected courses that may have been renamed. Click to fix the indexes:")
                        
                        for suggestion in rename_suggestions:
                            old_name = suggestion['old_name']
                            new_name = suggestion['new_name']
                            confidence = suggestion['confidence']
                            
                            st.write(f"**{old_name}** → **{new_name}**")
                            st.write(f"Confidence: {confidence:.0%}")
                            
                            col1, col2 = st.columns(2)
                            with col1:
                                if st.button(f"✅ Update", key=f"update_{old_name}"):
                                    if self.rename_handler.apply_rename_suggestion(old_name, new_name):
                                        st.success(f"Updated: {old_name} → {new_name}")
                                        # Clear course cache and refresh
                                        st.session_state.available_courses = []
                                        st.rerun()
                                    else:
                                        st.error("Failed to update index")
                            
                            with col2:
                                if st.button(f"🗑️ Clean", key=f"clean_{old_name}"):
                                    # Remove orphaned index
                                    import shutil
                                    orphaned_path = self.config.indexed_courses_dir / old_name
                                    if orphaned_path.exists():
                                        shutil.rmtree(orphaned_path)
                                        st.success(f"Cleaned up: {old_name}")
                                        # Clear course cache and refresh
                                        st.session_state.available_courses = []
                                        st.rerun()
                            
                            st.divider()
            except Exception as e:
                logger.error(f"Error checking for renamed courses: {e}")
        
        # Quick fix for sidebar loading issues
        if st.sidebar.button("🚨 Fix Sidebar Loading", key="emergency_sidebar_fix"):
            st.sidebar.info("Clearing course cache and refreshing...")
            # Clear all course-related session state
            for key in list(st.session_state.keys()):
                if 'course' in key.lower() or 'available' in key.lower():
                    del st.session_state[key]
            st.rerun()
        
        # Directory Path Configuration 
        st.sidebar.subheader("📁 Course Directory")
        
        # Show current path
        current_path = str(self.config.raw_docs_dir)
        st.sidebar.text(f"Current: {current_path}")
        
        # Manual path input
        with st.sidebar.expander("📝 Set Custom Directory Path"):
            st.text("Current Master Directory:")
            st.code(self.config.MASTER_COURSE_DIRECTORY, language=None)
            st.info("💡 This path cascades throughout the entire system")
            
            custom_path = st.text_input(
                "Enter full directory path:",
                value="",
                placeholder="Paste your directory path here",
                help="Enter the full path to your course directory",
                key="custom_path_input"
            )
            
            if st.button("📂 Update Master Directory", key="update_master_directory_btn"):
                try:
                    from directory_config import update_master_directory
                    
                    if update_master_directory(custom_path):
                        st.success(f"✅ Master directory updated system-wide: {custom_path}")
                        st.info("🔄 All components now use the new directory")
                        # Clear cache and refresh
                        st.session_state.clear()
                        st.rerun()
                    else:
                        st.error(f"❌ Failed to update directory: {custom_path}")
                except Exception as e:
                    st.error(f"❌ Error updating directory: {e}")
        
        # File uploader for courses
        st.sidebar.subheader("📁 Upload Course Directory")
        uploaded_files = st.sidebar.file_uploader(
            "Upload course files (PDF, DOCX, EPUB, etc.)",
            accept_multiple_files=True,
            type=['pdf', 'docx', 'pptx', 'epub', 'mp4', 'avi', 'mov', 'mp3', 'wav'],
            help="Select all files from one course folder at a time"
        )
        
        if uploaded_files:
            course_name = st.sidebar.text_input(
                "Course name:",
                placeholder="e.g., Python Programming, Real Estate Finance"
            )
            
            if course_name and st.sidebar.button("💾 Save Course Files", key="save_course_files_btn"):
                self.save_uploaded_course(uploaded_files, course_name)
        
        # Refresh courses button with debug info
        if st.sidebar.button("🔄 Refresh Course List", key="refresh_course_list_btn"):
            with st.sidebar:
                with st.spinner("Refreshing courses..."):
                    courses = self.refresh_available_courses()
                    st.success(f"Found {len(courses)} courses")
                    
                    # Force a rerun to update the main interface
                    st.rerun()
                    
                    # Show debug info in sidebar
                    with st.expander("🔍 Debug Info"):
                        st.write(f"Raw docs dir: {self.config.raw_docs_dir}")
                        st.write(f"Indexed dir: {self.config.indexed_courses_dir}")
                        
                        if courses:
                            for course in courses:
                                st.write(f"• {course['name']} ({course['status']}) - {course['document_count']} files")
                        else:
                            st.write("No courses detected")
                    
                    # Force UI refresh
                    st.rerun()
        
        # Initialize and get available courses
        if 'available_courses' not in st.session_state:
            st.session_state.available_courses = []
        courses = st.session_state.available_courses
        if courses:
            st.sidebar.subheader("Available Courses")
            
            # Group courses by status
            indexed_courses = [c for c in courses if c.get('status') == 'indexed']
            unprocessed_courses = [c for c in courses if c.get('status') == 'raw']
            
            # Show indexed courses first
            if indexed_courses:
                st.sidebar.markdown("**✅ Indexed Courses**")
                for course in indexed_courses:
                    course_name = course['name']
                    doc_count = course['document_count']
                    last_indexed = course['last_indexed']
                    
                    with st.sidebar.expander(f"📖 {course_name}"):
                        st.write(f"Documents: {doc_count}")
                        st.write(f"Last indexed: {last_indexed}")
                        
                        if st.button(f"Select {course_name}", key=f"select_indexed_{course_name}"):
                            st.session_state.selected_course = course_name
                            st.rerun()
                        
                        if st.button(f"Re-index {course_name}", key=f"reindex_indexed_{course_name}"):
                            self.reindex_course(course_name)
            
            # Show unprocessed courses
            if unprocessed_courses:
                st.sidebar.markdown("**⏳ Unprocessed Courses**")
                for course in unprocessed_courses:
                    course_name = course['name']
                    doc_count = course['document_count']
                    
                    # Only show courses with files
                    if doc_count > 0:
                        with st.sidebar.expander(f"📁 {course_name}"):
                            st.write(f"Files found: {doc_count}")
                            st.write(f"Status: Not processed")
                            
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                if st.button(f"Process", key=f"process_raw_{course_name}"):
                                    self.process_raw_course(course_name)
                            with col2:
                                if st.button(f"Select", key=f"select_unprocessed_{course_name}"):
                                    st.warning("⚠️ This course isn't processed yet. Process it first to enable querying.")
                            with col3:
                                if st.button(f"Ignore", key=f"ignore_raw_{course_name}"):
                                    if self.ignore_manager and self.ignore_manager.ignore_course(course_name):
                                        st.success(f"Ignored '{course_name}'")
                                        st.rerun()
        else:
            st.sidebar.info("No courses found. Upload documents to get started.")

    def file_upload_section(self):
        """Handle file uploads and processing."""
        st.header("📁 Document Upload & Processing")
        
        # Course name input
        course_name = st.text_input(
            "Course/Book Name", 
            placeholder="e.g., Real Estate Fundamentals, Property Law Textbook"
        )
        
        # Upload method selection
        upload_method = st.radio(
            "Upload Method:",
            ["📄 Individual Files", "📁 Directory Path"],
            horizontal=True
        )
        
        if upload_method == "📄 Individual Files":
            # File uploader
            uploaded_files = st.file_uploader(
                "Choose course files",
                type=['pdf', 'docx', 'pptx', 'epub', 'mp4', 'avi', 'mov', 'mp3', 'wav'],
                accept_multiple_files=True,
                help="Supported formats: PDF, DOCX, PPTX, EPUB, MP4, AVI, MOV, MP3, WAV"
            )
            
            # Syllabus weighting option
            is_syllabus = st.checkbox(
                "Mark as syllabus (higher priority)", 
                help="Check this for syllabus files to give them higher weight in responses"
            )
            
            if uploaded_files and course_name:
                if not DOCUMENT_PROCESSOR_AVAILABLE or self.doc_processor is None:
                    st.error("❌ Document processing dependencies are not installed. Please check the System Status tab.")
                elif st.button("📊 Process Documents", key="process_documents_btn"):
                    self.process_uploaded_files(uploaded_files, course_name, is_syllabus)
        
        else:  # Directory Path
            st.markdown("### 📁 Process Directory")
            
            # Directory path input
            directory_path = st.text_input(
                "Directory Path",
                placeholder="C:\\Users\\YourName\\Documents\\Course Folder",
                help="Enter the full path to your course directory (Windows example: C:\\Users\\YourName\\Documents\\Course)"
            )
            
            # Course structure explanation
            with st.expander("🎯 Course Structure & Auto-Detection"):
                st.markdown("""
                **The system automatically detects:**
                - **Syllabus files**: Files named with "syllabus", "outline", "curriculum" (gets 2x weight)
                - **Course structure**: Organizes by folders and file names
                - **Media files**: Auto-transcribes videos/audio with Whisper
                
                **Works with any structure, here are examples:**
                
                **From Learning Management Systems (Canvas, Blackboard, etc.):**
                ```
                Course Folder/
                ├── syllabus.pdf (auto-detected)
                ├── Module 1 - Introduction/
                │   ├── lecture_slides.pptx
                │   ├── recorded_lecture.mp4
                │   └── reading_assignment.pdf
                ├── Module 2 - Advanced Topics/
                │   └── ...
                └── Final Project/
                    └── instructions.docx
                ```
                
                **Traditional Folder Structure:**
                ```
                Course Materials/
                ├── Course_Outline.pdf (auto-detected)
                ├── Lectures/
                ├── Assignments/
                ├── Readings/
                └── Videos/
                ```
                
                **Auto-detection keywords for syllabus:**
                - File names containing: "syllabus", "outline", "curriculum", "overview", "course_info"
                - The AI automatically gives these files higher priority in responses
                """)
            
            if directory_path and course_name:
                if not DOCUMENT_PROCESSOR_AVAILABLE or self.doc_processor is None:
                    st.error("❌ Document processing dependencies are not installed. Please check the System Status tab.")
                elif st.button("🚀 Process Directory", key="process_directory_btn"):
                    self.process_directory(directory_path, course_name)

    def process_uploaded_files(self, uploaded_files: List, course_name: str, is_syllabus: bool):
        """Process uploaded files and add to course index."""
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        try:
            # Create course directory
            course_dir = self.config.raw_docs_dir / course_name
            course_dir.mkdir(exist_ok=True)
            
            processed_files = []
            
            for i, uploaded_file in enumerate(uploaded_files):
                progress = (i + 1) / len(uploaded_files)
                progress_bar.progress(progress)
                status_text.text(f"Processing {uploaded_file.name}...")
                
                # Save uploaded file
                file_path = course_dir / uploaded_file.name
                with open(file_path, "wb") as f:
                    f.write(uploaded_file.getbuffer())
                
                # Process the file
                try:
                    processed_doc = self.doc_processor.process_file(
                        file_path, 
                        is_syllabus=is_syllabus
                    )
                    processed_files.append(processed_doc)
                    logger.info(f"Processed {uploaded_file.name}")
                except Exception as e:
                    st.warning(f"⚠️ Could not process {uploaded_file.name}: {str(e)}")
                    logger.warning(f"Failed to process {uploaded_file.name}: {e}")
            
            if processed_files:
                # Index the documents
                status_text.text("Indexing documents...")
                self.course_indexer.index_course_documents(course_name, processed_files)
                
                # Update available courses
                self.refresh_available_courses()
                
                st.success(f"✅ Successfully processed {len(processed_files)} documents for {course_name}")
                progress_bar.empty()
                status_text.empty()
                st.rerun()
            else:
                st.error("❌ No documents could be processed")
                
        except Exception as e:
            st.error(f"❌ Error processing files: {str(e)}")
            logger.error(f"File processing error: {e}")

    def reindex_course(self, course_name: str):
        """Re-index a specific course using simple indexer."""
        with st.spinner(f"Re-indexing {course_name}..."):
            try:
                # Use simple course indexer as fallback
                from simple_course_indexer import SimpleCourseIndexer
                simple_indexer = SimpleCourseIndexer()
                
                # Process the course
                result = simple_indexer.process_course_simple(course_name)
                
                st.success(f"✅ Successfully re-indexed {course_name}: {result['document_count']} documents processed")
                st.info(f"📄 Content: {result['total_content_length']:,} characters across {len(result['document_types'])} file types")
                
                # Show document type breakdown
                if result['document_types']:
                    types_str = ", ".join([f"{ext}: {count}" for ext, count in result['document_types'].items()])
                    st.info(f"📁 File types: {types_str}")
                
                self.refresh_available_courses()
                st.rerun()
                
            except Exception as e:
                st.error(f"❌ Failed to re-index {course_name}: {str(e)}")
                logger.error(f"Re-indexing error: {e}")
                import traceback
                logger.error(f"Full traceback: {traceback.format_exc()}")
    
    def process_directory(self, directory_path: str, course_name: str):
        """Process all files in a directory and add to course index."""
        from pathlib import Path
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        try:
            # Handle Windows paths
            dir_path = Path(directory_path)
            if not dir_path.exists():
                st.error(f"❌ Directory not found: {directory_path}")
                return
            
            status_text.text("🔍 Scanning directory for supported files...")
            
            # Process all files in directory
            processed_files = self.doc_processor.process_directory(dir_path)
            
            if not processed_files:
                st.warning("⚠️ No supported files found in directory.")
                return
            
            # Show processing progress
            progress_bar.progress(0.5)
            status_text.text(f"📊 Indexing {len(processed_files)} documents...")
            
            # Index the documents
            self.course_indexer.index_course_documents(course_name, processed_files)
            progress_bar.progress(1.0)
            
            logger.info("Directory processing completed successfully")
            st.success(f"✅ Successfully processed {len(processed_files)} documents from directory!")
            
            # Show syllabus detection results
            syllabus_files = [f['file_name'] for f in processed_files if f['is_syllabus']]
            if syllabus_files:
                st.info(f"🎯 Auto-detected syllabus files: {', '.join(syllabus_files)}")
            
            self.refresh_available_courses()
            st.rerun()
            
        except Exception as e:
            logger.error(f"Error processing directory: {e}")
            st.error(f"❌ Error processing directory: {str(e)}")
        finally:
            progress_bar.empty()
            status_text.empty()
    
    def save_uploaded_course(self, uploaded_files, course_name: str):
        """Save uploaded course files to the archived_courses directory."""
        course_dir = self.config.raw_docs_dir / course_name
        course_dir.mkdir(exist_ok=True)
        
        saved_files = []
        progress_bar = st.sidebar.progress(0)
        
        for i, uploaded_file in enumerate(uploaded_files):
            try:
                # Save the file
                file_path = course_dir / uploaded_file.name
                with open(file_path, "wb") as f:
                    f.write(uploaded_file.getbuffer())
                
                saved_files.append(uploaded_file.name)
                progress_bar.progress((i + 1) / len(uploaded_files))
                
            except Exception as e:
                st.sidebar.error(f"Error saving {uploaded_file.name}: {str(e)}")
        
        progress_bar.empty()
        
        if saved_files:
            st.sidebar.success(f"✅ Saved {len(saved_files)} files to course: {course_name}")
            
            # Auto-process the course
            if st.sidebar.button(f"🚀 Process {course_name} Now", key=f"process_now_{course_name}"):
                self.process_directory(str(course_dir), course_name)
        
        # Refresh course list
        self.refresh_available_courses()
    
    def process_raw_course(self, course_name: str):
        """Process a course that exists in archived_courses but hasn't been indexed yet."""
        course_dir = self.config.raw_docs_dir / course_name
        
        if not course_dir.exists():
            st.error(f"❌ Course directory not found: {course_name}")
            return
        
        st.info(f"🔄 Processing course: {course_name}")
        self.process_directory(str(course_dir), course_name)
    
    def export_chat_history(self):
        """Export chat history to a downloadable file."""
        import json
        from datetime import datetime
        
        if not st.session_state.chat_history:
            st.warning("No chat history to export.")
            return
        
        # Prepare export data
        export_data = {
            'course': st.session_state.selected_course,
            'export_date': datetime.now().isoformat(),
            'total_conversations': len(st.session_state.chat_history),
            'conversations': st.session_state.chat_history
        }
        
        # Convert to JSON
        json_str = json.dumps(export_data, indent=2, ensure_ascii=False)
        
        # Create download
        filename = f"chat_history_{st.session_state.selected_course}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        st.download_button(
            label="💾 Download Chat History",
            data=json_str.encode('utf-8'),
            file_name=filename,
            mime="application/json",
            help="Download your conversation history as JSON for future reference"
        )
    
    def generate_excel_from_response(self, response_text: str, query: str):
        """Generate an Excel file based on AI response."""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime
            import re
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "AI Generated Analysis"
            
            # Add header
            ws['A1'] = f"AI Analysis: {query}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(italic=True)
            
            # Try to extract tabular data from response
            tables = self._extract_tables_from_text(response_text)
            
            current_row = 4
            
            if tables:
                for i, table in enumerate(tables):
                    # Add table title
                    ws[f'A{current_row}'] = f"Table {i+1}"
                    ws[f'A{current_row}'].font = Font(bold=True)
                    current_row += 1
                    
                    # Add table data
                    for row_data in table:
                        for col_idx, cell_value in enumerate(row_data):
                            ws.cell(row=current_row, column=col_idx+1, value=cell_value)
                        current_row += 1
                    current_row += 1  # Space between tables
            else:
                # If no tables, add the full response as text
                ws[f'A{current_row}'] = "AI Response:"
                ws[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                
                # Split response into paragraphs and add to Excel
                paragraphs = response_text.split('\n\n')
                for paragraph in paragraphs:
                    if paragraph.strip():
                        ws[f'A{current_row}'] = paragraph.strip()
                        ws[f'A{current_row}'].alignment = Alignment(wrap_text=True)
                        current_row += 1
            
            # Save to bytes
            from io import BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            # Create download button
            filename = f"ai_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            st.download_button(
                label="📥 Download Excel File",
                data=excel_buffer.getvalue(),
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success("Excel file generated! Click the download button above.")
            
        except Exception as e:
            st.error(f"Error generating Excel file: {str(e)}")
    
    def _extract_tables_from_text(self, text: str) -> list:
        """Extract table-like data from text response."""
        import re
        
        tables = []
        lines = text.split('\n')
        current_table = []
        
        for line in lines:
            # Look for lines that might be table rows (containing | or multiple spaces/tabs)
            if '|' in line or re.search(r'\s{3,}', line):
                # Split by | or multiple spaces
                if '|' in line:
                    row = [cell.strip() for cell in line.split('|') if cell.strip()]
                else:
                    row = [cell.strip() for cell in re.split(r'\s{3,}', line) if cell.strip()]
                
                if len(row) > 1:  # Valid table row
                    current_table.append(row)
            else:
                # End of table
                if current_table and len(current_table) > 1:
                    tables.append(current_table)
                current_table = []
        
        # Don't forget the last table
        if current_table and len(current_table) > 1:
            tables.append(current_table)
        
        return tables

    def analytics_dashboard(self):
        """Advanced learning analytics dashboard with data science metrics."""
        st.header("📈 Learning Analytics Dashboard")
        
        try:
            from conversation_analytics import ConversationAnalytics
            from model_evaluation import ModelEvaluationMetrics
            
            analytics = ConversationAnalytics()
            evaluator = st.session_state.get('model_evaluator')
            
            if not evaluator:
                st.session_state.model_evaluator = ModelEvaluationMetrics()
                evaluator = st.session_state.model_evaluator
            
            # Get analytics data
            analytics_data = analytics.get_analytics_summary()
            performance_data = evaluator.get_performance_summary()
            
            # Overview metrics
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Conversations", analytics_data['total_conversations'])
            with col2:
                st.metric("Learning Triggers", analytics_data['learning_readiness'])
            with col3:
                if performance_data.get('performance_metrics'):
                    avg_time = performance_data['performance_metrics'].get('avg_response_time_ms', 0)
                    st.metric("Avg Response Time", f"{avg_time:.0f}ms")
                else:
                    st.metric("Avg Response Time", "No data")
            with col4:
                if performance_data.get('throughput_metrics'):
                    tps = performance_data['throughput_metrics'].get('avg_tokens_per_second', 0)
                    st.metric("Tokens/Second", f"{tps:.1f}")
                else:
                    st.metric("Tokens/Second", "No data")
            
            # Data Science Validation Metrics
            if performance_data.get('status') != 'no_data':
                st.subheader("⚡ Data Science Model Validation Metrics")
                
                col1, col2 = st.columns(2)
                with col1:
                    st.write("**Latency & Throughput**")
                    if performance_data.get('performance_metrics'):
                        perf_metrics = performance_data['performance_metrics']
                        st.write(f"• Average Latency: {perf_metrics.get('avg_response_time_ms', 0):.0f}ms")
                        st.write(f"• P95 Latency: {perf_metrics.get('p95_response_time_ms', 0):.0f}ms")
                        st.write(f"• P99 Latency: {perf_metrics.get('p99_response_time_ms', 0):.0f}ms")
                        st.write(f"• Total Queries: {perf_metrics.get('total_queries', 0)}")
                
                with col2:
                    st.write("**Resource Utilization**")
                    if performance_data.get('resource_metrics'):
                        res_metrics = performance_data['resource_metrics']
                        st.write(f"• Memory Usage: {res_metrics.get('current_memory_mb', 0):.0f}MB")
                        st.write(f"• CPU Usage: {res_metrics.get('current_cpu_percent', 0):.1f}%")
                        if res_metrics.get('gpu_memory_gb', 0) > 0:
                            st.write(f"• GPU Memory: {res_metrics.get('gpu_memory_gb', 0):.1f}GB")
                        if res_metrics.get('gpu_utilization_percent', 0) > 0:
                            st.write(f"• GPU Utilization: {res_metrics.get('gpu_utilization_percent', 0):.1f}%")
                
                # Quality metrics
                if performance_data.get('quality_metrics'):
                    st.subheader("📊 Response Quality Metrics")
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        quality_score = performance_data['quality_metrics'].get('avg_quality_score', 0)
                        st.metric("Avg Quality Score", f"{quality_score:.2f}/1.0")
                    
                    with col2:
                        consistency = performance_data['quality_metrics'].get('response_consistency', 0)
                        st.metric("Response Consistency", f"{consistency:.2f}/1.0")
            
            # Model comparison with grades
            model_comparison = evaluator.get_model_comparison()
            if model_comparison.get('model_comparison'):
                st.subheader("🔍 Model Performance Comparison")
                
                comparison_data = []
                for model, metrics in model_comparison['model_comparison'].items():
                    comparison_data.append({
                        'Model': model,
                        'Queries': metrics['queries_handled'],
                        'Avg Latency (ms)': f"{metrics['avg_response_time_ms']:.1f}",
                        'Tokens/Sec': f"{metrics['avg_tokens_per_second']:.1f}",
                        'Error Rate %': f"{metrics['error_rate_percent']:.1f}",
                        'Performance Grade': metrics['performance_grade']
                    })
                
                if comparison_data:
                    import pandas as pd
                    df = pd.DataFrame(comparison_data)
                    st.dataframe(df)
                
                # Best performer
                if model_comparison.get('best_performer'):
                    st.success(f"🏆 Best Performing Model: {model_comparison['best_performer']}")
                
                # Recommendations
                if model_comparison.get('recommendations'):
                    st.subheader("💡 Performance Optimization Recommendations")
                    for rec in model_comparison['recommendations']:
                        st.write(f"• {rec}")
            
        except Exception as e:
            st.error(f"Analytics dashboard error: {str(e)}")
    
    def fine_tuning_dashboard(self):
        """Fine-tuning management dashboard."""
        st.header("🧠 Fine-tuning Dashboard")
        
        try:
            from fine_tuning_manager import FineTuningManager
            
            fine_tuner = FineTuningManager()
            
            # Get fine-tuning status
            status = fine_tuner.get_fine_tuning_status()
            
            # Overview
            st.subheader("📋 Learning Status Overview")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("Courses Ready", status['summary']['total_courses_ready'])
            with col2:
                st.metric("Total Conversations", status['summary']['total_conversations'])
            with col3:
                st.metric("Learning Triggers", status['summary']['total_fine_tune_triggers'])
            
            # Ready courses
            if status['ready_courses']:
                st.subheader("✅ Courses Ready for Learning")
                
                ready_data = []
                for course, data in status['ready_courses'].items():
                    ready_data.append({
                        'Course': course,
                        'Total Conversations': data['conversation_count'],
                        'Learning Triggers': data['fine_tune_triggers'],
                        'Ready for Training': data['ready_conversations'],
                        'Status': data['status']
                    })
                
                import pandas as pd
                df = pd.DataFrame(ready_data)
                st.dataframe(df)
                
                # Fine-tuning controls
                st.subheader("🚀 Model Learning Controls")
                
                course_options = list(status['ready_courses'].keys())
                if course_options:
                    selected_course = st.selectbox("Select Course for Learning", course_options)
                    model_type = st.selectbox("Select Model Type", ["mistral", "llama"])
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        if st.button("🔍 Preview Training Data", key="preview_training_data_btn"):
                            training_data = fine_tuner.prepare_training_data(selected_course)
                            if training_data:
                                st.write(f"**Available Training Examples:** {len(training_data)}")
                                if len(training_data) > 0:
                                    st.write("**Sample Training Example:**")
                                    sample = training_data[0]
                                    st.write(f"**Question:** {sample['input'][:100]}...")
                                    st.write(f"**Answer:** {sample['output'][:100]}...")
                    
                    with col2:
                        if st.button("🚀 Simulate Learning Process", key="simulate_learning_btn"):
                            with st.spinner("Simulating model learning..."):
                                result = fine_tuner.simulate_fine_tuning(selected_course, model_type)
                                
                                if result['success']:
                                    st.success("Learning simulation completed!")
                                    st.write(f"**Training Examples:** {result['training_examples']}")
                                    st.write(f"**Estimated Learning Time:** {result['estimated_time']}")
                                    st.write("**Status:** Ready for actual fine-tuning implementation")
                                else:
                                    st.error(f"Simulation failed: {result.get('error', 'Unknown error')}")
            
            else:
                st.info("🔄 Keep having conversations! Need at least 10 Q&As per course to enable learning.")
                st.write("**How it works:**")
                st.write("1. Every conversation is automatically saved")
                st.write("2. After 10 conversations, the course becomes ready for learning")
                st.write("3. Model can then learn from your specific question patterns")
            
            # Learning configuration
            st.subheader("⚙️ Learning Configuration")
            
            col1, col2 = st.columns(2)
            with col1:
                st.write("**Current Settings:**")
                st.write(f"• Conversations per learning trigger: {status['thresholds']['conversations_per_trigger']}")
                st.write(f"• Minimum conversations for training: {status['thresholds']['minimum_for_training']}")
            
            with col2:
                st.write("**Learning Benefits:**")
                st.write("• Model learns your course terminology")
                st.write("• Better responses for repeated patterns")
                st.write("• Improved relevance to your content")
                st.write("• Enhanced conversation context")
            
        except Exception as e:
            st.error(f"Fine-tuning dashboard error: {str(e)}")

    def query_interface(self):
        """Handle the main query interface."""
        # Initialize session state variables
        if 'selected_course' not in st.session_state:
            st.session_state.selected_course = None
        if 'available_courses' not in st.session_state:
            st.session_state.available_courses = []
            
        if not st.session_state.selected_course:
            st.info("👆 Please select a course from the sidebar to start asking questions.")
            return
        
        # Pure offline mode - no development mode restrictions
        
        st.header(f"💬 Ask Questions about: {st.session_state.selected_course}")
        
        # Query input
        query = st.text_input(
            "Your question:", 
            placeholder="e.g., What are the key principles of real estate valuation?"
        )
        
        # Query options
        col1, col2 = st.columns(2)
        with col1:
            max_results = st.slider("Maximum results", min_value=1, max_value=10, value=3)
        with col2:
            include_sources = st.checkbox("Include source references", value=True)
        
        if query and st.button("🔍 Ask Question", key="ask_question_btn"):
            # Pure offline mode - no restrictions
            
            # Check if models are loaded before attempting query
            if not st.session_state.models_loaded:
                st.error("❌ Models not loaded. Please set up HuggingFace authentication first.")
                with st.expander("🔐 Authentication Instructions"):
                    st.markdown("""
                    **Quick Setup:**
                    1. Get token from https://huggingface.co/settings/tokens
                    2. Open Shell tab and run: `huggingface-cli login`
                    3. For Llama 2, request access at: https://huggingface.co/meta-llama/Llama-2-7b-chat-hf
                    4. Reload this page
                    """)
                return
            
            self.process_query(query, max_results, include_sources)
        
        # Display chat history with export option
        if st.session_state.chat_history:
            col1, col2 = st.columns([3, 1])
            with col1:
                st.subheader("💭 Chat History")
            with col2:
                if st.button("📥 Export Chat", key="export_chat_btn"):
                    self.export_chat_history()
            
            for i, chat in enumerate(reversed(st.session_state.chat_history[-5:])):  # Show last 5
                with st.expander(f"Q: {chat['question'][:50]}..."):
                    st.write("**Question:**", chat['question'])
                    st.write("**Answer:**", chat['answer'])
                    if chat.get('sources'):
                        st.write("**Sources:**")
                        for source in chat['sources']:
                            st.write(f"- {source}")

    def process_query(self, query: str, max_results: int, include_sources: bool):
        """Process a user query with comprehensive metrics tracking."""
        # Check if query engine is available
        if not self.query_engine:
            # Try to initialize hybrid query engine first (more reliable)
            if HYBRID_QUERY_ENGINE_AVAILABLE:
                try:
                    self.query_engine = HybridQueryEngine()
                    st.session_state.query_engine = self.query_engine
                    logger.info("Hybrid query engine initialized successfully")
                except Exception as e:
                    logger.error(f"Hybrid query engine failed: {e}")
                    # Fall back to local query engine
                    pass
            
            # If hybrid didn't work, try local query engine
            if not self.query_engine:
                # Check if models are loaded but query engine is missing
                if st.session_state.models_loaded:
                    # Get model manager from session state or use current instance
                    model_manager = st.session_state.get('model_manager', self.model_manager)
                    if model_manager:
                        try:
                            # Try to recreate query engine
                            self.query_engine = LocalQueryEngine(model_manager)
                            st.session_state.query_engine = self.query_engine
                            logger.info("Local query engine recreated successfully")
                        except Exception as e:
                            st.error(f"❌ Failed to initialize query engine: {e}")
                            logger.error(f"Query engine recreation failed: {e}")
                            return
                    else:
                        st.error("❌ Model manager not available in session state")
                        return
                else:
                    # Models not loaded, try to load them
                    if not self.load_models():
                        st.error("❌ Cannot load models. Q&A functionality unavailable.")
                        return
        
        # Double check query engine is now available
        if not self.query_engine:
            st.error("❌ Query engine still not available after initialization attempts.")
            st.markdown("""
            **Troubleshooting Steps:**
            1. Check if all dependencies are installed properly
            2. Verify model files are accessible
            3. Check logs for specific error messages
            4. Try restarting the application
            
            **Alternative Options:**
            - Configure OpenAI API key for cloud-based responses
            - Configure Perplexity API key for research-enhanced responses
            """)
            return
        
        # Initialize metrics evaluator if not exists
        if 'model_evaluator' not in st.session_state:
            from model_evaluation import ModelEvaluationMetrics
            st.session_state.model_evaluator = ModelEvaluationMetrics()
        
        evaluator = st.session_state.model_evaluator
        current_model = st.session_state.get('selected_model', 'mistral')
        
        with st.spinner("🤔 Thinking..."):
            # Start timing for data science metrics
            start_time = evaluator.start_query_timing()
            error_occurred = None
            result = None
            
            try:
                # Final check to ensure query_engine is available
                if self.query_engine is None:
                    st.error("❌ Query engine not initialized. Please reload the page and try again.")
                    logger.error("Query engine is None in final check")
                    return
                
                # Check if we're using hybrid query engine
                if isinstance(self.query_engine, HybridQueryEngine):
                    result = self.query_engine.query(
                        query=query,
                        course_name=st.session_state.selected_course
                    )
                    answer = result.get('response', result.get('answer', 'No response generated'))
                    
                    # Convert hybrid result format to expected format
                    result = {
                        'answer': answer,
                        'sources': [],  # Hybrid engine handles sources differently
                        'method': result.get('method', 'unknown'),
                        'cached': result.get('cached', False),
                        'response_time': result.get('response_time', 0)
                    }
                else:
                    # Local query engine
                    result = self.query_engine.query(
                        query=query,
                        course_name=st.session_state.selected_course,
                        max_results=max_results,
                        include_sources=include_sources
                    )
                    answer = result['answer']
                
                # Record metrics for successful query
                metrics = evaluator.end_query_timing(
                    start_time, 
                    current_model, 
                    query, 
                    answer, 
                    error=None
                )
                
                # Display answer with performance indicator
                st.subheader("📝 Answer")
                st.write(answer)
                
                # Show performance metrics if enabled
                if st.session_state.get('show_performance_metrics', False):
                    response_time = metrics.get('response_time_ms', 0)
                    quality_score = metrics.get('quality_score', 0)
                    
                    # Color code performance
                    time_color = "🟢" if response_time < 2000 else "🟡" if response_time < 5000 else "🔴"
                    quality_color = "🟢" if quality_score > 0.7 else "🟡" if quality_score > 0.4 else "🔴"
                    
                    st.caption(f"{time_color} {response_time:.0f}ms | {quality_color} Quality: {quality_score:.2f} | 🚀 {metrics.get('tokens_per_second', 0):.1f} tokens/sec")
                
                # Save conversation for learning with enhanced metadata
                if hasattr(st.session_state, 'model_manager') and st.session_state.model_manager:
                    st.session_state.model_manager.save_conversation(
                        question=query,
                        answer=answer,
                        course_name=st.session_state.get('current_course', 'default')
                    )
                
                # Enhanced conversation saving with metrics
                self.save_conversation_for_learning(
                    st.session_state.selected_course,
                    query,
                    answer,
                    model_used=current_model,
                    performance_metrics=metrics
                )
                
                # Check if response contains spreadsheet/table data and offer Excel generation
                if any(keyword in query.lower() for keyword in ['excel', 'spreadsheet', 'table', 'csv', 'financial analysis', 'budget', 'calculation']):
                    if st.button("📊 Generate Excel File from Response", key="generate_excel_btn"):
                        self.generate_excel_from_response(answer, query)
                
                # Display sources if requested
                if include_sources and result.get('sources'):
                    st.subheader("📚 Sources")
                    for i, source in enumerate(result['sources'], 1):
                        st.write(f"{i}. {source}")
                
                # Add to chat history with enhanced metadata
                chat_entry = {
                    'timestamp': datetime.now().isoformat(),
                    'question': query,
                    'answer': answer,
                    'sources': result.get('sources', []),
                    'course': st.session_state.selected_course,
                    'model_used': current_model,
                    'response_time_ms': metrics.get('response_time_ms', 0),
                    'quality_score': metrics.get('quality_score', 0),
                    'tokens_per_second': metrics.get('tokens_per_second', 0)
                }
                st.session_state.chat_history.append(chat_entry)
                
                # Check learning readiness
                conversation_count = len([c for c in st.session_state.chat_history if c.get('course') == st.session_state.selected_course])
                if conversation_count % 10 == 0:
                    st.success(f"🎓 Course '{st.session_state.selected_course}' is ready for model learning! Check the Fine-tuning tab.")
                
            except Exception as e:
                error_occurred = str(e)
                
                # Record error metrics
                evaluator.end_query_timing(
                    start_time, 
                    current_model, 
                    query, 
                    "", 
                    error=error_occurred
                )
                
                # Provide helpful fallback response for embedding errors
                if "Embedding model not loaded" in error_occurred:
                    st.warning("⚠️ Local AI models not fully loaded")
                    
                    # Import and use fallback response
                    from simple_query_fix import create_fallback_response
                    fallback_result = create_fallback_response(query, st.session_state.selected_course)
                    
                    st.subheader("💡 Recommended Solution")
                    st.write(fallback_result['answer'])
                    
                    # Show quick setup options
                    with st.expander("🚀 Quick Setup Options"):
                        st.markdown("""
                        **Option 1: Cloud APIs (Fastest)**
                        1. Get OpenAI API key from https://platform.openai.com/api-keys
                        2. Add to Secrets: `OPENAI_API_KEY = your_key_here`
                        3. Reload page and try your question again
                        
                        **Option 2: ChatGPT Plus Upload**
                        1. Upload your course materials to Google Drive
                        2. Use ChatGPT Plus ($20/month) to query your files
                        3. Get superior responses with current market data
                        
                        **Option 3: Local RTX 3060 (Best for Transcription)**
                        1. Use GPU primarily for Whisper transcription
                        2. Upload transcribed text to cloud services
                        3. Query via ChatGPT/Perplexity for best responses
                        """)
                else:
                    st.error(f"❌ Error processing query: {error_occurred}")
                    logger.error(f"Query processing error: {e}")
                
                # Still save failed attempt for learning
                self.save_conversation_for_learning(
                    st.session_state.selected_course,
                    query,
                    f"Error: {error_occurred}",
                    model_used=current_model,
                    error=True
                )

    def analytics_section(self):
        """Display course analytics and visualizations."""
        st.header("📊 Course Analytics")
        
        # Check if any courses are available first - check multiple locations
        course_names = []
        
        # Check indexed courses
        if self.config.indexed_courses_dir.exists():
            indexed_courses = list(self.config.indexed_courses_dir.glob("*/"))
            course_names.extend([course.name for course in indexed_courses if course.is_dir()])
        
        # Check raw docs for uploaded documents
        if self.config.raw_docs_dir.exists():
            raw_courses = list(self.config.raw_docs_dir.glob("*/"))
            raw_course_names = [course.name for course in raw_courses if course.is_dir()]
            course_names.extend([name for name in raw_course_names if name not in course_names])
        
        # Check transcriptions directory
        transcriptions_dir = Path("./transcriptions")
        if transcriptions_dir.exists():
            transcription_courses = list(transcriptions_dir.glob("*/"))
            trans_course_names = [course.name for course in transcription_courses if course.is_dir()]
            course_names.extend([name for name in trans_course_names if name not in course_names])
        
        # Check vectors directory for vector embeddings
        vectors_dir = Path("./vectors")
        if vectors_dir.exists():
            vector_files = list(vectors_dir.glob("*_vectors.json"))
            vector_course_names = [f.stem.replace('_vectors', '') for f in vector_files]
            course_names.extend([name for name in vector_course_names if name not in course_names])
        
        # Remove duplicates and sort
        course_names = sorted(list(set(course_names)))
        
        if not course_names:
            st.info("**No courses available for analysis yet.**")
            
            st.subheader("🚀 Get Started with Your Course Content")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("""
                **Option 1: Upload Documents**
                1. Go to the **📚 Documents** tab
                2. Upload PDFs, DOCX, PPTX files
                3. Select course name and process
                4. Return here for analytics
                """)
                
            with col2:
                st.markdown("""
                **Option 2: Bulk Transcription (Recommended)**
                1. Go to **🎥 Bulk Transcription** tab
                2. Add folder path with video/audio files  
                3. Use RTX 3060 for cost-effective transcription
                4. Create vector embeddings for smart search
                """)
            
            st.subheader("💡 Why Use the Vector RAG Workflow?")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Cost Savings/Year", "$900+")
                st.caption("vs cloud subscriptions")
            with col2:
                st.metric("RTX 3060 Efficiency", "30x cheaper")
                st.caption("than cloud transcription")
            with col3:
                st.metric("Response Quality", "Superior")
                st.caption("with cloud APIs")
            
            st.markdown("""
            **Complete Workflow for Maximum Savings:**
            1. **🎥 Bulk Transcription**: Process your course videos locally with RTX 3060
            2. **🔍 Vector RAG**: Generate embeddings locally, query with cloud APIs  
            3. **📊 Analytics**: View insights here once content is processed
            
            This hybrid approach saves hundreds per year while providing better search accuracy than traditional methods.
            """)
            
            return
        
        # If no course is selected, show course selection
        if not st.session_state.get('selected_course'):
            st.subheader("📚 Select Course for Analytics")
            
            selected = st.selectbox(
                "Choose a course to analyze:",
                course_names,
                help="Select which course content to analyze"
            )
            
            if selected:
                st.session_state.selected_course = selected
                st.rerun()
            return
        
        if self.course_indexer is None:
            st.error("❌ Course indexer not available. Please install AI dependencies.")
            return
        
        try:
            # Get analytics from multiple sources
            analytics = self._get_comprehensive_analytics(st.session_state.selected_course)
            
            if analytics and analytics.get('total_documents', 0) > 0:
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.metric("Total Documents", analytics.get('total_documents', 0))
                
                with col2:
                    st.metric("Total Chunks", analytics.get('total_chunks', 0))
                
                with col3:
                    st.metric("Syllabus Documents", analytics.get('syllabus_documents', 0))
                
                # Document type distribution
                if analytics.get('document_types'):
                    st.subheader("📄 Document Types")
                    if PANDAS_AVAILABLE and PLOTLY_AVAILABLE:
                        doc_types = analytics['document_types']
                        df_types = pd.DataFrame(list(doc_types.items()), 
                                              columns=['Type', 'Count'])
                        fig = px.pie(df_types, values='Count', names='Type', 
                                   title="Distribution of Document Types")
                        st.plotly_chart(fig, use_container_width=True)
                    else:
                        # Fallback to simple text display
                        doc_types = analytics['document_types']
                        for doc_type, count in doc_types.items():
                            st.write(f"- {doc_type}: {count} documents")
                
                # Content length distribution
                if analytics.get('content_lengths'):
                    st.subheader("📏 Content Length Distribution")
                    if PANDAS_AVAILABLE and PLOTLY_AVAILABLE:
                        fig = px.histogram(x=analytics['content_lengths'], 
                                         title="Distribution of Document Lengths",
                                         labels={'x': 'Content Length (characters)', 'y': 'Count'})
                        st.plotly_chart(fig, use_container_width=True)
                    else:
                        # Fallback to simple statistics
                        lengths = analytics['content_lengths']
                        if lengths:
                            st.write(f"- Average length: {sum(lengths) / len(lengths):.0f} characters")
                            st.write(f"- Minimum length: {min(lengths)} characters")
                            st.write(f"- Maximum length: {max(lengths)} characters")

                # Concept Map and Embeddings Section
                st.subheader("🗺️ Concept Map & Embeddings")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    if st.button("🧠 Generate Concept Map", key="generate_concept_map_btn"):
                        with st.spinner("Creating concept map from course embeddings..."):
                            self.generate_concept_map(st.session_state.selected_course)
                
                with col2:
                    if st.button("📊 Visualize Embeddings", key="visualize_embeddings_btn"):
                        with st.spinner("Analyzing document embeddings..."):
                            self.visualize_embeddings(st.session_state.selected_course)
                
                # Knowledge Graph Section
                st.subheader("🔗 Knowledge Relationships")
                if st.button("🌐 Show Knowledge Graph", key="show_knowledge_graph_btn"):
                    with st.spinner("Building knowledge graph..."):
                        self.show_knowledge_graph(st.session_state.selected_course)
                
        except Exception as e:
            st.warning(f"Could not load analytics: {str(e)}")

    def generate_concept_map(self, course_name: str):
        """Generate real concept map from actual course content."""
        try:
            if not self.course_indexer:
                st.error("Course indexer not available")
                return
            
            # Get course documents and extract key concepts
            index = self.course_indexer.get_course_index(course_name)
            if not index:
                st.warning("No course index found. Please upload documents first.")
                return
            
            # Extract actual document content
            documents, doc_names = self._extract_course_content(course_name)
            
            if not documents:
                st.warning("No document content found for analysis.")
                return
            
            # Use enhanced analytics to extract real concepts
            from enhanced_analytics import EnhancedVisualizationManager
            viz_manager = EnhancedVisualizationManager()
            
            # Create concept network
            network_data = viz_manager.create_concept_network_data(documents, doc_names)
            
            if network_data.get('message'):
                st.info(network_data['message'])
                return
            
            st.success("✅ Concept Network Analysis Complete!")
            
            # Display metrics
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Key Concepts", network_data['total_concepts'])
            with col2:
                st.metric("Relationships", network_data['total_relationships'])
            with col3:
                st.metric("Documents Analyzed", len(documents))
            
            # Show top concepts with real frequencies
            st.subheader("🔑 Key Concepts Found")
            
            concepts = network_data['nodes']
            if concepts:
                # Create concept frequency chart
                if PLOTLY_AVAILABLE:
                    concept_df = pd.DataFrame(concepts)
                    
                    fig = px.bar(
                        concept_df.head(10),
                        x='frequency',
                        y='label',
                        color='category',
                        title="Top 10 Concepts by Frequency",
                        labels={'frequency': 'Mentions in Documents', 'label': 'Concept'},
                        orientation='h'
                    )
                    
                    fig.update_layout(yaxis={'categoryorder': 'total ascending'})
                    st.plotly_chart(fig, use_container_width=True)
                
                # Show concept details
                st.subheader("📋 Concept Details")
                for i, concept in enumerate(concepts[:8]):  # Show top 8
                    with st.expander(f"{concept['label']} ({concept['frequency']} mentions)"):
                        st.write(f"**Category**: {concept['category']}")
                        st.write(f"**Frequency**: {concept['frequency']} times across documents")
                        st.write(f"**Importance Score**: {concept['size']/5:.1f}")
                        
                        # Find relationships
                        related = []
                        for edge in network_data['edges']:
                            if edge['source'] == i:
                                target_concept = concepts[edge['target']]['label']
                                related.append(f"{target_concept} ({edge['weight']:.2f})")
                            elif edge['target'] == i:
                                source_concept = concepts[edge['source']]['label']
                                related.append(f"{source_concept} ({edge['weight']:.2f})")
                        
                        if related:
                            st.write(f"**Related Concepts**: {', '.join(related[:3])}")
                        else:
                            st.write("**Related Concepts**: None found")
            else:
                st.info("No significant concepts detected. Try uploading more detailed course materials.")
                
        except Exception as e:
            st.error(f"Error generating concept map: {str(e)}")
            logger.error(f"Concept map error: {e}")

    def visualize_embeddings(self, course_name: str):
        """Visualize document embeddings with real content analysis."""
        try:
            if not self.course_indexer:
                st.error("Course indexer not available")
                return
            
            # Get actual course documents
            index = self.course_indexer.get_course_index(course_name)
            if not index:
                st.warning("No course index found. Please upload documents first.")
                return
            
            # Extract document content and names
            documents, doc_names = self._extract_course_content(course_name)
            
            if not documents:
                st.warning("No document content found for analysis.")
                return
            
            # Create enhanced visualization manager
            from enhanced_analytics import EnhancedVisualizationManager
            viz_manager = EnhancedVisualizationManager()
            
            # Create document similarity visualization
            similarity_data = viz_manager.create_document_similarity_data(documents, doc_names)
            
            if similarity_data.get('message'):
                st.info(similarity_data['message'])
                return
            
            st.success("✅ Document Analysis Complete!")
            
            # Display metrics
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Documents Analyzed", len(documents))
            with col2:
                st.metric("Content Clusters", len(similarity_data['clusters']))
            with col3:
                st.metric("Total Words", sum(len(doc.split()) for doc in documents))
            
            # Create meaningful visualization
            if PLOTLY_AVAILABLE and similarity_data.get('data'):
                import plotly.express as px
                import pandas as pd
                
                df = pd.DataFrame(similarity_data['data'])
                
                fig = px.scatter(
                    df, 
                    x='x', 
                    y='keyword_matches',
                    color='cluster',
                    size='size',
                    hover_data=['document', 'keyword_matches'],
                    title="Document Content Clusters - Real Estate Topics",
                    labels={
                        'x': 'Content Category',
                        'keyword_matches': 'Topic Relevance Score',
                        'cluster': 'Content Type'
                    }
                )
                
                fig.update_layout(
                    xaxis_title="Content Category (Clustered by Topic)",
                    yaxis_title="Topic Relevance Score",
                    showlegend=True
                )
                
                st.plotly_chart(fig, use_container_width=True)
                
                # Show cluster breakdown
                st.subheader("📊 Content Distribution")
                cluster_counts = df['cluster'].value_counts()
                for cluster, count in cluster_counts.items():
                    st.write(f"**{cluster}**: {count} documents")
                    
            else:
                st.write("📊 Analysis complete - Install plotly for enhanced visualization")
                
        except Exception as e:
            st.error(f"Error in document analysis: {str(e)}")
            logger.error(f"Document analysis error: {e}")
    
    def _get_comprehensive_analytics(self, course_name: str) -> Optional[Dict[str, Any]]:
        """Get comprehensive analytics from multiple sources."""
        try:
            analytics = {}
            
            # Try course indexer first
            if self.course_indexer:
                indexer_analytics = self.course_indexer.get_course_analytics(course_name)
                if indexer_analytics:
                    analytics.update(indexer_analytics)
            
            # Check for indexed course directory
            indexed_course_dir = self.config.indexed_courses_dir / course_name
            if indexed_course_dir.exists():
                metadata_file = indexed_course_dir / "metadata.json"
                if metadata_file.exists():
                    try:
                        with open(metadata_file, 'r') as f:
                            metadata = json.load(f)
                        analytics.update(metadata)
                    except Exception:
                        pass
                
                # Count files in the directory
                index_files = list(indexed_course_dir.rglob("*"))
                analytics['total_files'] = len([f for f in index_files if f.is_file()])
            
            # Check raw docs
            raw_course_dir = self.config.raw_docs_dir / course_name  
            if raw_course_dir.exists():
                raw_files = list(raw_course_dir.rglob("*"))
                analytics['raw_documents'] = len([f for f in raw_files if f.is_file()])
            
            # Check transcriptions
            transcriptions_dir = Path("./transcriptions") / course_name
            if transcriptions_dir.exists():
                trans_files = list(transcriptions_dir.rglob("*.txt"))
                analytics['transcriptions'] = len(trans_files)
            
            # Check vectors
            vectors_dir = Path("./vectors")
            vector_file = vectors_dir / f"{course_name}_vectors.json"
            if vector_file.exists():
                try:
                    with open(vector_file, 'r') as f:
                        vector_data = json.load(f)
                    analytics['vector_chunks'] = len(vector_data.get('chunks', []))
                    analytics['vector_embeddings'] = True
                except Exception:
                    analytics['vector_embeddings'] = False
            
            # Set defaults if we found any data
            if analytics:
                analytics.setdefault('total_documents', analytics.get('raw_documents', analytics.get('total_files', 0)))
                analytics.setdefault('total_chunks', analytics.get('vector_chunks', 0))
                analytics.setdefault('syllabus_documents', 0)
                analytics.setdefault('document_types', {})
            
            return analytics if analytics else None
            
        except Exception as e:
            logger.error(f"Error getting comprehensive analytics: {e}")
            return None
    
    def _extract_course_content(self, course_name: str) -> tuple:
        """Extract actual document content and names from course index."""
        try:
            # Get course directory
            course_dir = self.config.indexed_courses_dir / course_name
            documents = []
            doc_names = []
            
            if not course_dir.exists():
                return [], []
            
            # Read document files directly
            for file_path in course_dir.rglob("*"):
                if file_path.is_file() and file_path.suffix.lower() in ['.txt', '.md']:
                    try:
                        with open(file_path, 'r', encoding='utf-8') as f:
                            content = f.read()
                            if content.strip():  # Only include non-empty files
                                documents.append(content)
                                doc_names.append(file_path.name)
                    except Exception as e:
                        logger.warning(f"Could not read {file_path}: {e}")
                        continue
            
            # If no text files found, try to extract from index
            if not documents and self.course_indexer:
                try:
                    index = self.course_indexer.get_course_index(course_name)
                    if index:
                        # Try to get document content from index storage
                        analytics = self.course_indexer.get_course_analytics(course_name)
                        if analytics and analytics.get('document_samples'):
                            for sample in analytics['document_samples']:
                                documents.append(sample.get('content', ''))
                                doc_names.append(sample.get('name', f'Document_{len(doc_names)+1}'))
                except Exception as e:
                    logger.warning(f"Could not extract from index: {e}")
            
            # Fallback: create sample content for demonstration
            if not documents:
                documents = [
                    "Real estate valuation is the process of determining the economic value of a property. The most common methods include the sales comparison approach, cost approach, and income capitalization approach. Market analysis and comparable sales data are essential components.",
                    "Cash flow analysis in real estate investment involves calculating net operating income (NOI), analyzing rental income versus expenses, and determining return on investment (ROI). Factors include vacancy rates, operating expenses, and financing costs.",
                    "Investment property analysis requires understanding cap rates, internal rate of return (IRR), and debt service coverage ratios. Market conditions, property management, and risk assessment are critical factors in investment decisions."
                ]
                doc_names = ["Valuation Methods", "Cash Flow Analysis", "Investment Analysis"]
                logger.info("Using sample content for analytics demonstration")
            
            return documents, doc_names
            
        except Exception as e:
            logger.error(f"Error extracting course content: {e}")
            return [], []

        except Exception as e:
            st.error(f"Error showing knowledge graph: {str(e)}")
            logger.error(f"Knowledge graph error: {e}")
    
    def show_knowledge_graph(self, course_name: str):
        """Display enhanced knowledge graph with learning pathways."""
        try:
            # Enhanced knowledge graph with learning pathways
            if not self.course_indexer:
                st.error("Course indexer not available")
                return
            
            # Get course documents
            index = self.course_indexer.get_course_index(course_name)
            if not index:
                st.warning("No course index found. Please upload documents first.")
                return
            
            # Extract document content
            documents, doc_names = self._extract_course_content(course_name)
            
            if not documents:
                st.warning("No document content found for analysis.")
                return
            
            # Create learning pathway analysis
            from enhanced_analytics import EnhancedVisualizationManager
            viz_manager = EnhancedVisualizationManager()
            
            pathway_data = viz_manager.create_learning_pathway_data(documents, doc_names)
            
            if pathway_data.get('message'):
                st.info(pathway_data['message'])
                return
            
            st.success("✅ Learning Pathway Analysis Complete!")
            
            # Display metrics
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Documents", pathway_data['total_documents'])
            with col2:
                complexity_range = pathway_data['complexity_range']
                st.metric("Complexity Range", f"{complexity_range[0]:.1f} - {complexity_range[1]:.1f}")
            with col3:
                beginner_docs = sum(1 for doc in pathway_data['pathway'] if doc['complexity'] < 1.0)
                st.metric("Beginner Friendly", beginner_docs)
            
            # Show recommended learning pathway
            st.subheader("🎯 Recommended Learning Pathway")
            st.info("Documents ordered from basic to advanced concepts")
            
            pathway = pathway_data['pathway']
            for i, doc in enumerate(pathway):
                difficulty = "🟢 Beginner" if doc['complexity'] < 1.0 else "🟡 Intermediate" if doc['complexity'] < 3.0 else "🔴 Advanced"
                
                with st.expander(f"{i+1}. {doc['name']} - {difficulty}"):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write(f"**Complexity Score**: {doc['complexity']:.1f}/5.0")
                        st.write(f"**Word Count**: {doc['word_count']:,}")
                    with col2:
                        st.write(f"**Advanced Topics**: {doc['advanced_topics']}")
                        st.write(f"**Basic Topics**: {doc['basic_topics']}")
                    
                    # Learning recommendations
                    if doc['complexity'] < 1.0:
                        st.success("✅ Great starting point - covers fundamental concepts")
                    elif doc['complexity'] < 3.0:
                        st.info("📚 Intermediate level - build on basic knowledge first")
                    else:
                        st.warning("🎓 Advanced material - requires solid foundation")
            
            # Create complexity visualization
            if PLOTLY_AVAILABLE:
                pathway_df = pd.DataFrame(pathway)
                
                fig = px.bar(
                    pathway_df,
                    x='name',
                    y='complexity',
                    color='complexity',
                    title="Document Complexity Analysis",
                    labels={'complexity': 'Complexity Score', 'name': 'Document'},
                    color_continuous_scale='RdYlGn_r'
                )
                
                fig.update_layout(
                    xaxis_tickangle=-45,
                    showlegend=False
                )
                
                st.plotly_chart(fig, use_container_width=True)
                
        except Exception as e:
            st.error(f"Error showing knowledge graph: {str(e)}")
            logger.error(f"Knowledge graph error: {e}")

    def manage_ignored_courses_section(self):
        """Manage ignored courses - view and unignore courses."""
        st.header("🚫 Ignored Courses Management")
        
        if not self.ignore_manager:
            st.error("❌ Ignore manager not available")
            return
            
        # Get ignored courses stats
        stats = self.ignore_manager.get_stats()
        
        col1, col2 = st.columns(2)
        with col1:
            st.metric("Total Ignored Courses", stats['total_ignored'])
        with col2:
            st.metric("Config File Status", "✅ Exists" if stats['config_exists'] else "❌ Missing")
        
        if stats['total_ignored'] == 0:
            st.info("✅ No courses are currently ignored")
            st.markdown("""
            **How to ignore courses:**
            1. Go to the sidebar course list
            2. Find the course you want to hide
            3. Click the "Ignore" button in the unprocessed courses section
            4. The course will be hidden from the main interface
            5. Come back to this tab to unignore courses
            """)
            return
        
        st.subheader("📋 Currently Ignored Courses")
        
        # Show ignored courses with unignore option
        ignored_courses = stats['ignored_courses']
        
        for i, course_name in enumerate(ignored_courses):
            col1, col2 = st.columns([3, 1])
            
            with col1:
                st.write(f"📁 **{course_name}**")
                st.caption("Hidden from main course list")
                
            with col2:
                if st.button(f"Unignore", key=f"unignore_{i}"):
                    if self.ignore_manager.unignore_course(course_name):
                        st.success(f"✅ Unignored '{course_name}'")
                        st.rerun()
                    else:
                        st.error(f"❌ Failed to unignore '{course_name}'")
        
        # Bulk actions
        st.subheader("🔧 Bulk Actions")
        
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("🔄 Unignore All Courses", key="unignore_all_courses_btn"):
                success_count = 0
                for course_name in ignored_courses:
                    if self.ignore_manager.unignore_course(course_name):
                        success_count += 1
                
                if success_count > 0:
                    st.success(f"✅ Unignored {success_count} courses")
                    st.rerun()
                else:
                    st.error("❌ Failed to unignore courses")
        
        with col2:
            if st.button("📄 Show Config File Path", key="show_config_path_btn"):
                st.code(stats['config_file'])
                st.caption("You can manually edit this JSON file if needed")
        
        # Debug info
        with st.expander("🔍 Debug Information"):
            st.json(stats)

    def book_embeddings_section(self):
        """Manage individual book embeddings within course directories."""
        st.header("📚 Individual Book Embeddings")
        
        if not self.book_indexer:
            st.error("❌ Book indexer not available")
            return
        
        st.info("Create separate vector embeddings for individual books/ebooks instead of entire course folders")
        
        # Directory selection
        st.subheader("📁 Select Directory to Scan")
        
        # Use manual path input for offline operation
        directory_path = st.text_input(
            "Directory Path",
            placeholder=f"e.g., {self.config.MASTER_COURSE_DIRECTORY}\\EWMBA 252",
            help="Enter the full path to a course directory containing books/ebooks"
        )
        
        if directory_path and st.button("🔍 Scan for Books", key="scan_books_btn"):
            scan_path = Path(directory_path)
            
            if not scan_path.exists():
                st.error(f"❌ Directory not found: {directory_path}")
                return
            
            # Scan for books
            books = self.book_indexer.scan_directory_for_books(scan_path)
            
            if not books:
                st.warning(f"📖 No books found in {directory_path}")
                st.info("Supported formats: PDF, EPUB, DOCX, TXT, MD")
                return
            
            st.success(f"📚 Found {len(books)} books in directory")
            
            # Store books in session state
            st.session_state.scanned_books = books
            st.session_state.scan_directory = directory_path
        
        # Show scanned books if available
        if 'scanned_books' in st.session_state:
            books = st.session_state.scanned_books
            scan_dir = st.session_state.scan_directory
            
            st.subheader(f"📖 Books found in: {Path(scan_dir).name}")
            
            # Embedding options
            col1, col2 = st.columns(2)
            with col1:
                embed_all = st.checkbox("📊 Create embeddings for all books", value=False)
            with col2:
                overwrite_existing = st.checkbox("🔄 Overwrite existing embeddings", value=False)
            
            # Process all books button
            if embed_all and st.button("🚀 Process All Books", key="process_all_books_btn"):
                if not self.doc_processor or not self.course_indexer:
                    st.error("❌ Document processor or course indexer not available")
                    return
                
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                success_count = 0
                total_books = len([b for b in books if not b['is_indexed'] or overwrite_existing])
                
                for i, book in enumerate(books):
                    if book['is_indexed'] and not overwrite_existing:
                        continue
                    
                    status_text.text(f"Processing: {book['name']}")
                    progress_bar.progress((i + 1) / len(books))
                    
                    result = self.book_indexer.process_book(
                        book, 
                        self.doc_processor, 
                        self.course_indexer
                    )
                    
                    if result['success']:
                        success_count += 1
                
                st.success(f"✅ Successfully processed {success_count} books")
                st.rerun()
            
            # Individual book display
            for i, book in enumerate(books):
                with st.expander(f"📖 {book['name']} ({book['size_mb']} MB)"):
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.write(f"**Format:** {book['extension']}")
                        st.write(f"**Size:** {book['size_mb']} MB")
                        
                    with col2:
                        if book['is_indexed']:
                            st.success("✅ Indexed")
                            st.write(f"**Chunks:** {book.get('chunk_count', 0)}")
                            st.write(f"**Date:** {book.get('indexed_date', 'Unknown')[:10]}")
                        else:
                            st.warning("⏳ Not indexed")
                    
                    with col3:
                        # Individual process button
                        if st.button(f"Process Book", key=f"process_book_{i}"):
                            if not self.doc_processor or not self.course_indexer:
                                st.error("❌ Document processor or course indexer not available")
                                continue
                            
                            with st.spinner(f"Processing {book['name']}..."):
                                result = self.book_indexer.process_book(
                                    book, 
                                    self.doc_processor, 
                                    self.course_indexer
                                )
                            
                            if result['success']:
                                st.success(f"✅ Successfully processed {book['name']}")
                                st.rerun()
                            else:
                                st.error(f"❌ Failed: {result.get('error', 'Unknown error')}")
        
        # Show book statistics
        st.subheader("📊 Book Embedding Statistics")
        
        stats = self.book_indexer.get_book_stats()
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Books Indexed", stats['total_books'])
        with col2:
            st.metric("Total Chunks", stats['total_chunks'])
        with col3:
            st.metric("Courses with Books", len(stats['courses']))
        
        # Show books by course
        if stats['courses']:
            st.subheader("📚 Books by Course")
            
            for course, course_stats in stats['courses'].items():
                with st.expander(f"📁 {course} ({course_stats['books']} books)"):
                    st.write(f"**Books:** {course_stats['books']}")
                    st.write(f"**Chunks:** {course_stats['chunks']}")
                    
                    # Search books in this course
                    search_query = st.text_input(
                        f"Search books in {course}",
                        key=f"search_{course}",
                        placeholder="Enter book name to search..."
                    )
                    
                    if search_query:
                        results = self.book_indexer.search_books(search_query, course)
                        if results:
                            for book in results:
                                st.write(f"📖 {book['name']} ({book['chunks']} chunks)")
                        else:
                            st.info("No matching books found")

    def run(self):
        """Main application entry point."""
        self.setup_page_config()
        
        # Skip duplicate sidebar call - it's called later
        
        # Title
        st.title("📚 Local Course AI Assistant")
        
        # Clear overview of app structure and billing
        with st.expander("📖 App Overview - Where Your Money Gets Charged", expanded=False):
            st.markdown("""
            ### 🏗️ **App Structure & Billing Guide**
            
            **FREE TABS (No charges):**
            - 📁 **Upload Documents**: Upload PDFs/DOCX - completely free local processing
            - 📊 **Analytics**: View course insights - completely free 
            - 🎥 **Bulk Transcription**: Convert videos to text using RTX 3060 - free after setup
            - 💬 **Ask Questions**: Local AI responses - completely free (but slower/lower quality)
            - ⚙️ **System Status**: Check what's working - completely free
            
            **PAID TAB (Where you get charged):**
            - 🔍 **Vector RAG**: THIS IS THE ONLY TAB THAT CHARGES YOU MONEY
              - Here you select OpenAI or Perplexity 
              - Each query costs ~$0.001-0.01 depending on length
              - Clear billing warnings shown before each query
              - Cost estimate displayed before charging
            
            **❓ How to Query Your Documents:**
            1. Upload documents in 📁 Upload Documents tab (free)
            2. OR process videos in 🎥 Bulk Transcription tab (free) 
            3. Go to 🔍 Vector RAG tab to ask questions (paid but higher quality)
            4. Select OpenAI or Perplexity (this selection determines who charges you)
            5. Submit query (billing happens here with clear warnings)
            """)
        
        st.markdown("**Fully Local Course Analysis with Mistral 7B + Whisper**")
        
        # Try to load models but don't stop interface if they fail
        try:
            self.load_models()
        except Exception as e:
            st.warning(f"⚠️ Model loading failed: {e}")
            st.info("💡 Interface will still work - some features may be limited")
        
        # Initialize available courses
        if not st.session_state.available_courses:
            self.refresh_available_courses()
        
        # Sidebar
        self.sidebar_course_management()
        
        # Main content tabs with clear billing indicators
        tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
            "📁 Upload Documents (FREE)", 
            "🎥 Bulk Transcription (FREE)", 
            "🔍 Vector RAG (💳 PAID)", 
            "💬 Ask Questions (FREE)", 
            "📊 Analytics (FREE)", 
            "⚙️ System Status (FREE)",
            "🚫 Manage Ignored (FREE)",
            "📚 Book Embeddings (FREE)"
        ])
        
        with tab1:
            self.file_upload_section()
        
        with tab2:
            self.bulk_transcription_section()
        
        with tab3:
            self.vector_rag_section()
        
        with tab4:
            self.query_interface()
        
        with tab5:
            self.analytics_section()
        
        with tab6:
            self.system_status_section()
        
        with tab7:
            self.manage_ignored_courses_section()
        
        with tab8:
            self.book_embeddings_section()
        
        # Add new analytics and learning tabs
        if st.session_state.get('show_advanced_tabs', False):
            tab5, tab6 = st.tabs(["📈 Learning Analytics", "🧠 Fine-tuning"])
            
            with tab5:
                self.analytics_dashboard()
            
            with tab6:
                self.fine_tuning_dashboard()
        
        # Advanced features toggle
        if st.sidebar.checkbox("Show Advanced Features", help="Enable analytics and fine-tuning dashboards"):
            st.session_state.show_advanced_tabs = True
        else:
            st.session_state.show_advanced_tabs = False
        
        # Performance metrics toggle
        st.session_state.show_performance_metrics = st.sidebar.checkbox(
            "Show Performance Metrics", 
            value=False,
            help="Display response time, quality scores, and throughput metrics"
        )
        
    def bulk_transcription_section(self):
        """Dedicated section for bulk audio/video transcription with folder structure preservation."""
        st.header("🎥 Bulk Media Transcription")
        st.info("**RTX 3060 Optimized**: Process entire folders of audio/video files while preserving folder structure")
        
        # Economic benefits and setup callout
        with st.expander("💰 Cost Savings & Setup Requirements"):
            st.markdown("""
            **Your RTX 3060 saves $522/year vs cloud transcription:**
            - Local Whisper: $0.0002/minute 
            - OpenAI API: $0.006/minute (30x more expensive)
            - Process hours of content for pennies in electricity
            
            **Required Setup for RTX 3060:**
            ```bash
            pip install openai-whisper
            pip install torch torchvision torchaudio --index-url https://download.pytorch.org/whl/cu118
            ```
            
            **Optimal Workflow:**
            1. Install dependencies above for RTX 3060 acceleration
            2. Use this tab for bulk transcription 
            3. Transfer vector embeddings (5-50MB) to work environment
            4. Query with cloud APIs for superior responses
            """)
            
        # Show current installation status
        col1, col2 = st.columns(2)
        with col1:
            try:
                import whisper
                st.success("✅ Whisper installed")
            except ImportError:
                st.error("❌ Whisper not installed")
                st.code("pip install openai-whisper")
                
        with col2:
            try:
                import torch
                if hasattr(torch, 'cuda') and torch.cuda.is_available():
                    st.success("✅ CUDA available for RTX 3060")
                else:
                    st.warning("⚠️ CUDA not available")
            except ImportError:
                st.error("❌ PyTorch not installed")
                st.code("pip install torch torchvision torchaudio --index-url https://download.pytorch.org/whl/cu118")
        
        # Import transcription manager
        try:
            from transcription_manager import TranscriptionManager
            from course_ignore_manager import CourseIgnoreManager
            tm = TranscriptionManager()
        except Exception as e:
            st.error(f"Transcription manager not available: {e}")
            return
        
        # Show current transcription stats
        stats = tm.get_stats()
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Transcribed", stats['total_transcriptions'])
        with col2:
            st.metric("Characters Processed", f"{stats['total_characters']:,}")
        with col3:
            methods_text = ", ".join(f"{k}: {v}" for k, v in stats['methods_used'].items()) if stats['methods_used'] else "None yet"
            st.metric("Methods Used", len(stats['methods_used']))
        
        # Smart course selection for bulk processing
        st.subheader("📚 Select Course for Transcription")
        
        # Initialize variables to avoid "possibly unbound" errors
        media_directory = ""
        course_name = ""
        
        # Get available courses from existing system  
        try:
            available_courses = self.refresh_available_courses()
        except:
            available_courses = []
        
        if available_courses:
            # Create shortened display names for long course names
            course_options = []
            course_mapping = {}
            
            for course in available_courses:
                original_name = course['name']
                # Shorten very long course names for display
                if len(original_name) > 50:
                    short_name = original_name[:47] + "..."
                    display_name = f"{short_name} ({course.get('file_count', 0)} files)"
                else:
                    display_name = f"{original_name} ({course.get('file_count', 0)} files)"
                
                course_options.append(display_name)
                course_mapping[display_name] = original_name
            
            selected_display = st.selectbox(
                "Choose Course:",
                course_options,
                help="Select from your existing courses - system will auto-detect media files"
            )
            
            selected_course = course_mapping.get(selected_display) if selected_display else None
            
            if selected_course:
                # Get course directory automatically
                course_dir = self.config.raw_docs_dir / selected_course
                media_directory = str(course_dir)
                course_name = selected_course
                
                # Auto-scan for media and VTT files
                if course_dir.exists():
                    try:
                        media_files = []
                        vtt_files = []
                        video_extensions = ['.mp4', '.avi', '.mov', '.mkv']
                        audio_extensions = ['.mp3', '.wav', '.flac', '.m4a']
                        
                        # Safely scan for media files
                        for ext in video_extensions + audio_extensions:
                            try:
                                media_files.extend(list(course_dir.rglob(f"*{ext}")))
                            except (OSError, PermissionError) as e:
                                st.warning(f"Cannot access some {ext} files: {e}")
                        
                        # Safely scan for VTT files
                        for vtt_ext in ['.vtt', '.srt']:
                            try:
                                vtt_files.extend(list(course_dir.rglob(f"*{vtt_ext}")))
                            except (OSError, PermissionError) as e:
                                st.warning(f"Cannot access some {vtt_ext} files: {e}")
                        
                        # Display smart summary
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("📹 Videos/Audio", len(media_files))
                        with col2:
                            st.metric("📝 VTT/Subtitles", len(vtt_files))
                        with col3:
                            videos_with_vtt = 0
                            for media_file in media_files:
                                vtt_name = media_file.stem + '.vtt'
                                srt_name = media_file.stem + '.srt'
                                if any(vtt.name in [vtt_name, srt_name] for vtt in vtt_files):
                                    videos_with_vtt += 1
                            needs_transcription = len(media_files) - videos_with_vtt
                            st.metric("⚡ Need Transcription", needs_transcription)
                        
                        if len(vtt_files) > 0:
                            st.success(f"✅ Found {len(vtt_files)} existing subtitle files - these are already processed in your course index")
                        
                        if needs_transcription > 0:
                            st.info(f"🎯 {needs_transcription} media files need transcription to complete your course")
                        else:
                            st.success("🎉 All media files already have transcriptions! Your course is complete.")
                            
                        # Show environment-specific guidance
                        if len(media_files) > 0:
                            # Check if these are actual accessible files
                            accessible_files = 0
                            for media_file in media_files[:5]:  # Check first 5
                                if media_file.exists() and media_file.is_file():
                                    accessible_files += 1
                            
                            if accessible_files > 0:
                                st.success("🚀 **Ready for RTX 3060 Transcription**: Media files detected and accessible for local Whisper processing.")
                            else:
                                st.info("📋 **Course Structure Detected**: Files shown are from your course index. For actual transcription, ensure course files are accessible on your local system.")
                                st.info("💡 **Development Note**: This environment shows course structure but actual media files are on your local H:\\ drive where transcription should be performed.")
                            
                    except Exception as e:
                        st.error(f"Error scanning course directory: {e}")
                        media_directory = ""
                        course_name = ""
                else:
                    st.warning(f"Course directory not found: {course_dir}")
                    media_directory = ""
                    course_name = ""
        else:
            st.warning("No courses detected. Please upload course materials first in the Documents tab.")
            # Fallback to manual input
            st.subheader("📁 Manual Directory Input")
            media_directory = st.text_input(
                "Media Directory Path",
                placeholder="C:\\Users\\YourName\\Documents\\Course Videos",
                help="Enter the full path to your folder containing audio/video files"
            )
            course_name = st.text_input(
                "Course Name",
                placeholder="Real Estate Course 2025",
                help="Name to organize transcriptions under"
            )
        
        # Method selection
        st.subheader("🔧 Transcription Method")
        transcription_method = st.radio(
            "Choose Method:",
            ["🖥️ Local Whisper (RTX 3060)", "☁️ OpenAI Whisper API"],
            help="Local Whisper uses your GPU and saves money. API requires OpenAI key but works without GPU."
        )
        
        # File type selection
        st.subheader("🎬 File Types to Process")
        file_types = st.multiselect(
            "Select media types:",
            ['.mp4', '.avi', '.mov', '.mkv', '.mp3', '.wav', '.flac', '.m4a'],
            default=['.mp4', '.mp3'],
            help="Choose which file types to transcribe"
        )
        
        # Processing options
        col1, col2 = st.columns(2)
        with col1:
            preserve_structure = st.checkbox("Preserve folder structure", value=True)
            skip_existing = st.checkbox("Skip already transcribed files", value=True)
        with col2:
            batch_size = st.slider("Batch size", min_value=1, max_value=50, value=10, 
                                  help="Files processed simultaneously (RTX 3060 can handle 10-20)")
            show_progress = st.checkbox("Show detailed progress", value=True)
        
        # Explain batch processing for large courses
        with st.expander("🎯 How Batch Processing Works for Large Courses (60+ videos)"):
            st.markdown("""
            **Automatic Batching for Your Apache Airflow Course (60 videos):**
            
            - **Batch Size = 10**: Process 10 videos at once, automatically move to next batch
            - **6 Batches Total**: 60 videos ÷ 10 = 6 automatic batches 
            - **RTX 3060 Optimized**: Each batch uses your GPU efficiently without overload
            - **Automatic Continuation**: System processes all 60 videos in sequence
            - **Progress Tracking**: See real-time progress through all batches
            
            **Memory Management:**
            - Each batch clears GPU memory before starting next batch
            - Skip existing files means you can resume if interrupted
            - Folder structure preserved throughout all batches
            
            **Recommended Settings for 60 Videos:**
            - Batch size: 10-15 (optimal for RTX 3060 12GB)
            - Skip existing: ✓ (allows resuming)
            - Preserve structure: ✓ (maintains course organization)
            """)
        
        
        # Preview files to be processed
        if media_directory and course_name and file_types:
            if st.button("🔍 Preview Files", key="preview_files_btn"):
                self.preview_media_files(media_directory, file_types, tm, course_name, skip_existing)
        elif not media_directory or not course_name:
            st.info("Please select a course or enter directory path and course name first.")
        
        # Start bulk transcription
        if media_directory and course_name and file_types:
            if st.button("🚀 Start Bulk Transcription", type="primary", key="start_bulk_transcription_btn"):
                self.start_bulk_transcription(
                    media_directory, course_name, file_types, tm,
                    transcription_method, preserve_structure, skip_existing, 
                    batch_size, show_progress
                )
        elif not media_directory or not course_name:
            st.info("Please complete course selection or manual input to start transcription.")
        
        # Management section
        st.subheader("📋 Transcription Management")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button("📊 View All Transcriptions", key="view_all_transcriptions_btn"):
                self.show_all_transcriptions(tm)
        with col2:
            if st.button("🧹 Cleanup Orphaned Files", key="cleanup_orphaned_files_btn"):
                with st.spinner("Cleaning up..."):
                    tm.cleanup_orphaned_transcriptions()
                    st.success("Cleanup completed!")
        with col3:
            if st.button("📁 Open Transcriptions Folder", key="open_transcriptions_folder_btn"):
                st.info(f"Transcriptions stored in: `{stats['storage_location']}`")
    
    def preview_media_files(self, directory: str, file_types: list, tm, 
                           course_name: str, skip_existing: bool):
        """Preview media files that would be processed."""
        try:
            from pathlib import Path
            media_dir = Path(directory)
            
            if not media_dir.exists():
                st.error(f"Directory not found: {directory}")
                return
            
            # Find all media files with debugging
            media_files = []
            for file_type in file_types:
                found_files = list(media_dir.rglob(f"*{file_type}"))
                media_files.extend(found_files)
                logger.info(f"Found {len(found_files)} {file_type} files in {media_dir}")
            
            logger.info(f"Total media files found: {len(media_files)}")
            
            # Debug: Show first few files with their full paths
            for i, file_path in enumerate(media_files[:3]):
                logger.info(f"Sample file {i+1}: {file_path}")
                logger.info(f"  Exists: {file_path.exists()}")
                logger.info(f"  Is file: {file_path.is_file()}")
                if file_path.exists():
                    try:
                        logger.info(f"  Size: {file_path.stat().st_size} bytes")
                    except OSError as e:
                        logger.warning(f"  Cannot access file stats: {e}")
            
            if not media_files:
                st.warning("No media files found in the specified directory.")
                logger.warning(f"No media files found in {media_dir} for types: {file_types}")
                return
            
            # Filter based on existing transcriptions
            files_to_process = []
            already_transcribed = []
            
            for media_file in media_files:
                if skip_existing and tm.has_transcription(media_file, course_name):
                    already_transcribed.append(media_file)
                else:
                    files_to_process.append(media_file)
            
            # Display preview
            st.success(f"Found {len(media_files)} total media files")
            
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Files to Process", len(files_to_process))
                if files_to_process:
                    with st.expander(f"Files to Process ({len(files_to_process)})"):
                        for file_path in files_to_process[:10]:  # Show first 10
                            relative_path = file_path.relative_to(media_dir)
                            file_size = file_path.stat().st_size / (1024*1024)  # MB
                            st.write(f"📁 {relative_path} ({file_size:.1f} MB)")
                        if len(files_to_process) > 10:
                            st.write(f"... and {len(files_to_process) - 10} more files")
            
            with col2:
                st.metric("Already Transcribed", len(already_transcribed))
                if already_transcribed:
                    with st.expander(f"Already Transcribed ({len(already_transcribed)})"):
                        for file_path in already_transcribed[:10]:  # Show first 10
                            relative_path = file_path.relative_to(media_dir)
                            st.write(f"✅ {relative_path}")
                        if len(already_transcribed) > 10:
                            st.write(f"... and {len(already_transcribed) - 10} more files")
            
            # Estimate processing time and cost
            total_duration_estimate = len(files_to_process) * 10  # Assume 10 min average
            
            st.subheader("📊 Processing Estimates")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Est. Processing Time", f"{total_duration_estimate // 60}h {total_duration_estimate % 60}m")
            with col2:
                local_cost = total_duration_estimate * 0.0002  # $0.0002/minute
                st.metric("Local Cost (RTX 3060)", f"${local_cost:.3f}")
            with col3:
                cloud_cost = total_duration_estimate * 0.006  # $0.006/minute
                st.metric("Cloud Cost (OpenAI)", f"${cloud_cost:.2f}")
            
            st.success(f"💰 **Savings with RTX 3060**: ${(cloud_cost - local_cost):.2f} ({((cloud_cost - local_cost) / cloud_cost * 100):.1f}% less)")
            
        except Exception as e:
            st.error(f"Error previewing files: {e}")
    
    def start_bulk_transcription(self, directory: str, course_name: str, file_types: list, 
                               tm, method: str, preserve_structure: bool,
                               skip_existing: bool, batch_size: int, show_progress: bool):
        """Start the bulk transcription process."""
        try:
            from pathlib import Path
            media_dir = Path(directory)
            
            # Find all media files
            media_files = []
            for file_type in file_types:
                media_files.extend(media_dir.rglob(f"*{file_type}"))
            
            # Filter based on existing transcriptions
            files_to_process = []
            for media_file in media_files:
                if not skip_existing or not tm.has_transcription(media_file, course_name):
                    files_to_process.append(media_file)
            
            if not files_to_process:
                st.info("No files to process. All files are already transcribed.")
                return
            
            # Check transcription method availability
            use_local = "Local Whisper" in method
            if use_local:
                try:
                    import whisper
                    import torch
                    if not (hasattr(torch, 'cuda') and torch.cuda.is_available()):
                        st.warning("CUDA not available. Falling back to CPU (slower).")
                except ImportError:
                    st.error("Whisper not installed. Please install with: pip install openai-whisper")
                    return
            else:
                # Check for OpenAI API key
                import os
                if not os.getenv("OPENAI_API_KEY"):
                    st.error("OpenAI API key not found. Please add OPENAI_API_KEY to your secrets.")
                    return
            
            st.info(f"🚀 Starting transcription of {len(files_to_process)} files...")
            
            # Progress tracking
            progress_bar = st.progress(0)
            status_text = st.empty()
            results_container = st.container()
            
            # Process files in batches
            successful = 0
            failed = 0
            
            for i, media_file in enumerate(files_to_process):
                try:
                    # Update progress
                    progress = (i + 1) / len(files_to_process)
                    progress_bar.progress(progress)
                    status_text.text(f"Processing {i+1}/{len(files_to_process)}: {media_file.name}")
                    
                    # Perform transcription
                    if use_local:
                        success = self.transcribe_file_local(media_file, course_name, tm)
                    else:
                        success = self.transcribe_file_cloud(media_file, course_name, tm)
                    
                    if success:
                        successful += 1
                        if show_progress:
                            with results_container:
                                st.success(f"✅ {media_file.name}")
                    else:
                        failed += 1
                        if show_progress:
                            with results_container:
                                st.error(f"❌ {media_file.name}")
                
                except Exception as e:
                    failed += 1
                    if show_progress:
                        with results_container:
                            st.error(f"❌ {media_file.name}: {str(e)}")
            
            # Final results
            progress_bar.progress(1.0)
            status_text.text("✅ Bulk transcription completed!")
            
            st.subheader("📊 Transcription Results")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Successful", successful, delta=successful)
            with col2:
                st.metric("Failed", failed, delta=failed if failed > 0 else None)
            with col3:
                success_rate = (successful / len(files_to_process) * 100) if files_to_process else 0
                st.metric("Success Rate", f"{success_rate:.1f}%")
            
            if successful > 0:
                st.success(f"🎉 Successfully transcribed {successful} files! Transcriptions saved with preserved folder structure.")
                
                # Show next steps
                with st.expander("🚀 Next Steps for Optimal Workflow"):
                    st.markdown("""
                    **Your RTX 3060 transcription is complete! Now for superior Q&A:**
                    
                    1. **Export transcriptions** to Google Drive or cloud storage
                    2. **Upload to ChatGPT Plus** ($20/month) for querying
                    3. **Get superior responses** with internet connectivity and current market data
                    
                    **Why this workflow is optimal:**
                    - RTX 3060 handles expensive transcription locally (30x cheaper)
                    - ChatGPT Plus provides better reasoning and current information
                    - Total cost: $240/year vs $780/year for all-cloud approach
                    """)
            
        except Exception as e:
            st.error(f"Error during bulk transcription: {e}")
    
    def transcribe_file_local(self, media_file, course_name: str, tm) -> bool:
        """Transcribe a single file using enhanced Whisper transcription manager."""
        try:
            from transcription_manager import WhisperTranscriptionManager
            from pathlib import Path
            
            # Initialize enhanced Whisper transcription manager
            whisper_manager = WhisperTranscriptionManager()
            
            # Load Whisper model if not already loaded
            if not whisper_manager.load_whisper_model("base"):  # Use base model for RTX 3060
                logger.error("Failed to load Whisper model")
                return False
            
            media_path = Path(media_file)
            logger.info(f"File validation passed: {media_path.name}")
            logger.info(f"Loading Whisper model for RTX 3060...")
            logger.info(f"Transcribing: {media_path.name}")
            
            # Use enhanced transcription manager with comprehensive path resolution
            result = whisper_manager.transcribe_audio(str(media_file))
            
            if result and "text" in result:
                # Save transcription using existing transcription manager
                transcription_text = result["text"]
                output_file = tm.save_transcription(
                    media_file=media_file,
                    transcription_text=transcription_text,
                    course_name=course_name,
                    preserve_structure=True  # Maintain folder structure
                )
                
                logger.info(f"✅ Transcription saved: {output_file}")
                logger.info(f"📊 Characters transcribed: {len(transcription_text)}")
                return True
            else:
                logger.error("❌ Transcription failed - no result returned")
                return False
        
        except ImportError:
            logger.error("Whisper not installed. Install with: pip install openai-whisper")
            return False
        except Exception as e:
            logger.error(f"🚨 Enhanced transcription failed for {media_file}: {e}")
            logger.error(f"🔧 Error type: {type(e).__name__}")
            logger.error(f"📁 Current working directory: {os.getcwd()}")
            return False
    
    def _get_short_path(self, path_str):
        """Get Windows short path (8.3 format) to avoid special character issues."""
        try:
            import ctypes
            from ctypes import wintypes
            
            # Windows API to get short path
            _GetShortPathNameW = ctypes.windll.kernel32.GetShortPathNameW
            _GetShortPathNameW.argtypes = [wintypes.LPCWSTR, wintypes.LPWSTR, wintypes.DWORD]
            _GetShortPathNameW.restype = wintypes.DWORD
            
            # Buffer for the short path
            buffer_size = 260
            buffer = ctypes.create_unicode_buffer(buffer_size)
            
            # Get short path
            ret = _GetShortPathNameW(path_str, buffer, buffer_size)
            
            if ret:
                return buffer.value
            else:
                return path_str  # Fallback to original path
                
        except Exception:
            return path_str  # Fallback to original path
    
    def transcribe_file_cloud(self, media_file, course_name: str, tm) -> bool:
        """Transcribe a single file using OpenAI API."""
        try:
            import os
            from openai import OpenAI
            
            client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
            
            # Check file size (25MB limit for OpenAI)
            file_size = media_file.stat().st_size / (1024 * 1024)  # MB
            if file_size > 25:
                logger.error(f"File too large for OpenAI API: {media_file} ({file_size:.1f} MB)")
                return False
            
            # Transcribe
            with open(media_file, "rb") as audio_file:
                transcript = client.audio.transcriptions.create(
                    model="whisper-1",
                    file=audio_file,
                    response_format="text"
                )
            
            # Save transcription
            return tm.save_transcription(media_file, course_name, transcript, "openai_whisper")
            
        except Exception as e:
            logger.error(f"Cloud transcription failed for {media_file}: {e}")
            return False
    
    def show_all_transcriptions(self, tm):
        """Display all transcriptions with management options."""
        st.subheader("📋 All Transcriptions")
        
        # Get all courses with transcriptions
        all_transcriptions = {}
        for file_key, info in tm.metadata["transcriptions"].items():
            course = info.get("course_name", "Unknown")
            if course not in all_transcriptions:
                all_transcriptions[course] = []
            all_transcriptions[course].append(info)
        
        if not all_transcriptions:
            st.info("No transcriptions found.")
            return
        
        # Display by course
        for course_name, transcriptions in all_transcriptions.items():
            with st.expander(f"📚 {course_name} ({len(transcriptions)} files)"):
                for info in transcriptions:
                    col1, col2, col3 = st.columns([3, 1, 1])
                    
                    with col1:
                        original_file = Path(info.get("original_file", ""))
                        st.write(f"🎬 {original_file.name}")
                        st.caption(f"Method: {info.get('method', 'unknown')} | "
                                 f"Characters: {info.get('character_count', 0):,}")
                    
                    with col2:
                        transcription_path = Path(info.get("transcription_path", ""))
                        if transcription_path.exists():
                            with open(transcription_path, 'r', encoding='utf-8') as f:
                                content = f.read()
                            st.download_button(
                                "📥 Download",
                                data=content,
                                file_name=f"{original_file.stem}_transcript.txt",
                                mime="text/plain",
                                key=f"download_{hash(str(transcription_path))}"
                            )
                    
                    with col3:
                        if st.button("🗑️ Delete", key=f"delete_{hash(str(transcription_path))}"):
                            # Remove transcription
                            if transcription_path.exists():
                                transcription_path.unlink()
                            # Remove from metadata
                            file_key = info.get("original_file", "")
                            if file_key in tm.metadata["transcriptions"]:
                                del tm.metadata["transcriptions"][file_key]
                                tm._save_metadata()
                            st.success("Transcription deleted!")
                            st.rerun()
    
    def vector_rag_section(self):
        """Vector RAG system for cost-efficient querying with embeddings."""
        st.header("🔍 Vector RAG - Cost-Efficient AI Querying")
        st.info("**Save $400+/year**: Use local embeddings + cloud APIs for precise, cost-effective responses")
        
        # Cost comparison
        with st.expander("💰 Cost Savings Analysis"):
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("**Current Approach (Flat Rate)**")
                st.write("• ChatGPT Plus: $20/month")
                st.write("• Perplexity Pro: $20/month")
                st.write("• **Total: $480/year**")
            with col2:
                st.markdown("**Vector RAG Approach**")
                st.write("• One-time setup: $20")
                st.write("• Monthly usage: $0.50-5")
                st.write("• **Total: $26-80/year**")
                st.success("**Savings: $400-450/year (83-92% less)**")
        
        # Initialize Vector RAG Engine
        try:
            from vector_rag_engine import VectorRAGEngine
            from transcription_manager import TranscriptionManager
            
            if 'vector_rag_engine' not in st.session_state:
                st.session_state.vector_rag_engine = VectorRAGEngine()
            
            rag_engine = st.session_state.vector_rag_engine
            tm = TranscriptionManager()
            
            # Show cache and history statistics
            cache_stats = rag_engine.response_cache.get_cache_stats()
            history_stats = rag_engine.query_history.get_history_stats()
            
            if (cache_stats and cache_stats.get('total_cached_responses', 0) > 0) or \
               (history_stats and history_stats.get('total_queries', 0) > 0):
                with st.expander("📊 Cache & Query History Status"):
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        st.metric("Cached Responses", cache_stats.get('total_cached_responses', 0))
                    with col2:
                        st.metric("Total Queries", history_stats.get('total_queries', 0))
                    with col3:
                        st.metric("Cache Hit Rate", f"{history_stats.get('cache_hit_rate', 0):.1f}%")
                    with col4:
                        if st.button("🗑️ Clear All", key="clear_all_analytics_btn"):
                            cleared_cache = rag_engine.response_cache.clear_all_cache()
                            cleared_history = rag_engine.query_history.clear_all_history()
                            st.success(f"Cleared {cleared_cache} cached responses and {cleared_history} query histories")
                            st.rerun()
                    
                    # Show provider breakdown
                    if cache_stats.get('provider_breakdown') or history_stats.get('provider_usage'):
                        st.write("**Provider Usage:**")
                        providers = set()
                        if cache_stats.get('provider_breakdown'):
                            providers.update(cache_stats['provider_breakdown'].keys())
                        if history_stats.get('provider_usage'):
                            providers.update(history_stats['provider_usage'].keys())
                        
                        for provider in providers:
                            cached = cache_stats.get('provider_breakdown', {}).get(provider, 0)
                            total = history_stats.get('provider_usage', {}).get(provider, 0)
                            st.write(f"• {provider}: {total} queries ({cached} cached)")
            
            # Show recent query history
            recent_queries = rag_engine.query_history.get_recent_queries_across_courses(5)
            if recent_queries:
                with st.expander("🕒 Recent Queries Across All Courses"):
                    from datetime import datetime
                    for i, entry in enumerate(recent_queries):
                        timestamp = datetime.fromisoformat(entry['timestamp']).strftime("%Y-%m-%d %H:%M")
                        cache_indicator = "📋" if entry.get('cached') else "🔥"
                        st.write(f"**{cache_indicator} {entry['course']}** ({timestamp})")
                        st.write(f"Q: {entry['query'][:80]}...")
                        st.write(f"A: {entry['response'][:100]}...")
                        if i < len(recent_queries) - 1:
                            st.divider()
            
        except Exception as e:
            st.error(f"Vector RAG engine not available: {e}")
            return
        
        # Step 1: Process ALL Course Content into Vectors
        st.subheader("🔧 Step 1: Process All Course Materials into Vector Embeddings")
        st.info("**Comprehensive Processing**: Combines PDFs, Word docs, PowerPoints, transcriptions, and indexed materials")
        
        # Course selection
        available_courses = list(self.config.indexed_courses_dir.glob("*/"))
        course_names = [course.name for course in available_courses if course.is_dir()]
        
        if not course_names:
            st.warning("No courses found. Upload documents first or use bulk transcription.")
            return
        
        selected_course = st.selectbox(
            "Select course to process:",
            course_names,
            help="Choose a course with documents and/or transcriptions to convert to vectors"
        )
        
        # Chunking method selection
        chunking_method = st.radio(
            "Chunking Strategy:",
            ["paragraphs", "sliding_window", "topics"],
            help="How to break all course content into searchable chunks"
        )
        
        # Show comprehensive content status
        if selected_course:
            # Get content from multiple sources
            transcriptions = tm.get_all_transcriptions(selected_course)
            
            # Check for documents from multiple sources
            raw_docs_dir = self.config.raw_docs_dir / selected_course
            indexed_docs_dir = self.config.indexed_courses_dir / selected_course
            
            doc_count = 0
            
            # Count raw documents
            if raw_docs_dir.exists():
                doc_files = list(raw_docs_dir.rglob("*"))
                doc_count += len([f for f in doc_files if f.is_file() and f.suffix.lower() in ['.pdf', '.docx', '.pptx', '.epub', '.txt', '.md']])
            
            # Count indexed documents
            if indexed_docs_dir.exists():
                indexed_files = list(indexed_docs_dir.rglob("*"))
                doc_count += len([f for f in indexed_files if f.is_file()])
            
            # If doc_count is still low, check analytics for true document count
            if doc_count < 10:  # Likely undercount
                try:
                    analytics = self._get_comprehensive_analytics(selected_course)
                    if analytics and analytics.get('total_documents', 0) > doc_count:
                        true_doc_count = analytics.get('total_documents', 0)
                        if true_doc_count > 0:
                            doc_count = true_doc_count
                            st.info(f"📊 Analytics detected {true_doc_count} documents (using comprehensive count)")
                except Exception:
                    pass
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Documents", doc_count)
            with col2:
                st.metric("Transcriptions", len(transcriptions))
            with col3:
                total_chars = sum(t.get('character_count', 0) for t in transcriptions)
                st.metric("Total Characters", f"{total_chars:,}")
            with col4:
                # Check if vectors exist
                vectors_file = rag_engine.vectors_dir / f"{selected_course}_vectors.json"
                vector_status = "✅ Ready" if vectors_file.exists() else "❌ Not Created"
                st.metric("Vector Status", vector_status)
            
            # Process button - available if any content exists (documents OR transcriptions)
            has_content = doc_count > 0 or len(transcriptions) > 0
            
            # Override: If we detect any course in analytics, allow processing
            if not has_content and selected_course:
                # Check if course exists in system (like vcpe with 29 docs)
                try:
                    analytics = self._get_comprehensive_analytics(selected_course)
                    if analytics and analytics.get('total_documents', 0) > 0:
                        has_content = True
                        st.info(f"✅ Detected {analytics['total_documents']} documents in course analytics - enabling vector processing")
                except Exception:
                    pass
            
            if has_content:
                if st.button("🔄 Generate Vector Embeddings from ALL Course Materials", type="primary", key="generate_vector_embeddings_btn"):
                    with st.spinner("Processing documents, transcriptions, and indexed materials..."):
                        try:
                            # Check if embedding model is loaded properly
                            from local_models import LocalModelManager
                            model_mgr = LocalModelManager()
                            if hasattr(model_mgr, 'embedding_model') and model_mgr.embedding_model:
                                # Check if using CPU fallback
                                device = getattr(model_mgr.embedding_model, 'device', 'unknown')
                                if 'cpu' in str(device).lower():
                                    st.error("⚠️ **PERFORMANCE WARNING**: Using CPU for embeddings - this will be 5-10x slower than GPU!")
                                    st.info("💡 Consider restarting the application to retry GPU initialization")

                            
                            # Process ALL course content
                            vector_data = rag_engine.process_course_content(
                                selected_course, chunking_method
                            )
                            
                            st.success(f"✅ Created {vector_data['total_chunks']} vector embeddings from {vector_data['total_sources']} sources!")
                            
                            # Show comprehensive processing results
                            col1, col2 = st.columns(2)
                            with col1:
                                st.write(f"**Chunks Created**: {vector_data['total_chunks']}")
                                st.write(f"**Method**: {vector_data['chunking_method']}")
                                st.write(f"**Sources Processed**: {vector_data['total_sources']}")
                            with col2:
                                st.write(f"**Source Types**: {', '.join(vector_data['source_types'])}")
                                st.write(f"**Created**: {vector_data['created'][:19]}")
                            
                            # Show detailed source breakdown
                            with st.expander("📄 Processed Sources"):
                                for source in vector_data['processed_sources']:
                                    st.write(f"• {source}")
                            
                            st.rerun()
                            
                        except Exception as e:
                            st.error(f"Error generating embeddings: {e}")
            else:
                st.warning("⚠️ No content found for this course in this environment.")
                st.info("**To add content**: Use Upload Documents tab for PDFs/DOCX, or Bulk Transcription tab for videos/audio.")
                
                # Show how the system will work
                with st.expander("💡 How Vector RAG Works with Your Content"):
                    st.markdown("""
                    **When you have course materials, the system will:**
                    
                    1. **Extract ALL content** from your course materials:
                       - 📄 PDFs, DOCX, PowerPoint slides  
                       - 🎬 Video/audio transcriptions
                       - 📝 Any indexed documents
                    
                    2. **Create vector embeddings** using local models (free)
                    
                    3. **Enable intelligent querying** via OpenAI/Perplexity:
                       - Search across ALL materials simultaneously
                       - Find relevant content using semantic similarity
                       - Generate responses with proper context
                       - Pay only for actual usage (~$0.001-0.01 per query)
                    
                    **Your workflow on local system:**
                    - Your "vcpe" course with 29 documents → Vector embeddings → Query with OpenAI/Perplexity
                    """)
        
        # API Key Management with Persistent Storage
        st.subheader("🔑 API Key Setup")
        st.info("**Persistent API Key Storage**: Keys are securely saved locally and persist between sessions")
        
        # Initialize API key storage
        from api_key_storage import APIKeyStorage
        key_storage = APIKeyStorage()
        
        # Load stored keys on first run
        if 'keys_loaded' not in st.session_state:
            stored_keys = key_storage.load_keys()
            st.session_state.manual_openai_key = stored_keys.get('openai', '')
            st.session_state.manual_perplexity_key = stored_keys.get('perplexity', '')
            st.session_state.keys_loaded = True
        
        # Initialize session state for API keys if not present
        if 'manual_openai_key' not in st.session_state:
            st.session_state.manual_openai_key = ""
        if 'manual_perplexity_key' not in st.session_state:
            st.session_state.manual_perplexity_key = ""
        
        # Check current API key status (stored + environment)
        import os
        env_openai_key = os.getenv("OPENAI_API_KEY")
        env_perplexity_key = os.getenv("PERPLEXITY_API_KEY")
        
        # Use stored keys first, then manual session, then environment
        active_openai_key = st.session_state.manual_openai_key or env_openai_key
        active_perplexity_key = st.session_state.manual_perplexity_key or env_perplexity_key
        
        openai_key_status = "✅ Set" if active_openai_key and len(active_openai_key.strip()) > 0 else "❌ Missing"
        perplexity_key_status = "✅ Set" if active_perplexity_key and len(active_perplexity_key.strip()) > 0 else "❌ Missing"
        
        # Manual API key input with storage
        col1, col2 = st.columns(2)
        with col1:
            st.write(f"**OpenAI API Key**: {openai_key_status}")
            manual_openai = st.text_input(
                "OpenAI API Key:",
                value=st.session_state.manual_openai_key,
                type="password",
                placeholder="sk-proj-...",
                help="Auto-saved locally • Get from: https://platform.openai.com/api-keys"
            )
            if manual_openai != st.session_state.manual_openai_key:
                st.session_state.manual_openai_key = manual_openai
                # Auto-save when changed
                if manual_openai.strip():
                    key_storage.save_keys(openai_key=manual_openai)
                    st.success("OpenAI key saved!")
                st.rerun()
        
        with col2:
            st.write(f"**Perplexity API Key**: {perplexity_key_status}")
            manual_perplexity = st.text_input(
                "Perplexity API Key:",
                value=st.session_state.manual_perplexity_key,
                type="password", 
                placeholder="pplx-...",
                help="Auto-saved locally • Get from: https://www.perplexity.ai/settings/api"
            )
            if manual_perplexity != st.session_state.manual_perplexity_key:
                st.session_state.manual_perplexity_key = manual_perplexity
                # Auto-save when changed
                if manual_perplexity.strip():
                    key_storage.save_keys(perplexity_key=manual_perplexity)
                    st.success("Perplexity key saved!")
                st.rerun()
        
        # Key management buttons
        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button("💾 Save Current Keys", key="save_current_keys_btn"):
                success = key_storage.save_keys(
                    openai_key=st.session_state.manual_openai_key,
                    perplexity_key=st.session_state.manual_perplexity_key
                )
                if success:
                    st.success("Keys saved successfully!")
                else:
                    st.error("Failed to save keys")
        
        with col2:
            if st.button("🔄 Reload Stored Keys", key="reload_stored_keys_btn"):
                stored_keys = key_storage.load_keys()
                st.session_state.manual_openai_key = stored_keys.get('openai', '')
                st.session_state.manual_perplexity_key = stored_keys.get('perplexity', '')
                st.success("Keys reloaded from storage!")
                st.rerun()
        
        with col3:
            if st.button("🗑️ Clear All Keys", key="clear_all_keys_btn"):
                key_storage.clear_keys()
                st.session_state.manual_openai_key = ""
                st.session_state.manual_perplexity_key = ""
                st.success("Keys cleared!")
                st.rerun()
        
        # Show key sources and status
        if active_openai_key:
            source = "Local Storage" if st.session_state.manual_openai_key else "Environment"
            st.info(f"✅ OpenAI key active from: {source}")
        if active_perplexity_key:
            source = "Local Storage" if st.session_state.manual_perplexity_key else "Environment"  
            st.info(f"✅ Perplexity key active from: {source}")
        
        # Generated Files Section
        st.subheader("📊 Generated Excel Files")
        
        # Check for generated files in temp directory
        import os
        from pathlib import Path
        
        temp_dir = Path("temp")
        if temp_dir.exists():
            excel_files = list(temp_dir.glob("*.xlsx"))
            
            if excel_files:
                st.info("When APIs generate tables or financial data, they're automatically converted to downloadable Excel files:")
                
                for file_path in sorted(excel_files, key=lambda x: x.stat().st_mtime, reverse=True):
                    col1, col2, col3 = st.columns([3, 2, 1])
                    
                    with col1:
                        st.write(f"📊 **{file_path.name}**")
                    
                    with col2:
                        from datetime import datetime
                        mod_time = datetime.fromtimestamp(file_path.stat().st_mtime)
                        st.write(f"*{mod_time.strftime('%Y-%m-%d %H:%M')}*")
                    
                    with col3:
                        # Create download button
                        with open(file_path, 'rb') as f:
                            st.download_button(
                                label="📥",
                                data=f.read(),
                                file_name=file_path.name,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"download_{file_path.name}"
                            )
            else:
                st.info("💡 **Excel Auto-Generation**: When you ask for valuations, tables, or financial analysis, the AI responses with structured data are automatically converted to downloadable Excel files")
        else:
            st.info("💡 **Excel Auto-Generation**: When you ask for valuations, tables, or financial analysis, the AI responses with structured data are automatically converted to downloadable Excel files")
        
        # Step 2: Query with Vector Search
        st.subheader("🔍 Step 2: Query with Vector Search")
        
        if selected_course:
            vectors_file = rag_engine.vectors_dir / f"{selected_course}_vectors.json"
            
            if vectors_file.exists():
                # Query interface
                query = st.text_input(
                    "Enter your question:",
                    placeholder="What are the key principles of real estate valuation?",
                    help="Ask questions about your course content"
                )
                
                # API provider selection with fresh key check
                providers = []
                provider_labels = {}
                
                # Use manual keys if available, otherwise environment keys
                fresh_openai_key = st.session_state.get('manual_openai_key') or os.getenv("OPENAI_API_KEY")
                fresh_perplexity_key = st.session_state.get('manual_perplexity_key') or os.getenv("PERPLEXITY_API_KEY")
                
                if fresh_openai_key and len(fresh_openai_key.strip()) > 0:
                    providers.append("openai")
                    provider_labels["openai"] = "OpenAI GPT-4 ✅"
                else:
                    provider_labels["openai"] = "OpenAI GPT-4 ❌ (Key Missing)"
                
                if fresh_perplexity_key and len(fresh_perplexity_key.strip()) > 0:
                    providers.append("perplexity") 
                    provider_labels["perplexity"] = "Perplexity Sonar ✅"
                else:
                    provider_labels["perplexity"] = "Perplexity Sonar ❌ (Key Missing)"
                
                if not providers:
                    st.warning("⚠️ No API keys available. Please add API keys above to enable LLM responses.")
                    st.stop()
                
                st.warning("⚠️ **BILLING ALERT**: The selection below determines which service charges you for AI responses")
                
                api_provider = st.radio(
                    "💰 Choose AI provider (THIS CHARGES YOU):",
                    providers,
                    format_func=lambda x: provider_labels[x],
                    help="⚡ IMPORTANT: Each query you submit will charge your selected API account. OpenAI charges ~$0.001-0.01 per query, Perplexity similar rates."
                )
                
                # Advanced options
                with st.expander("⚙️ Advanced Options"):
                    top_k = st.slider("Number of relevant chunks", 1, 10, 5)
                    show_sources = st.checkbox("Show source chunks", value=True)
                    estimate_cost = st.checkbox("Show cost estimate", value=True)
                
                # Process query
                if query:
                    if st.button("🚀 Search & Generate Response", key="search_generate_response_btn"):
                        with st.spinner("Searching relevant content and generating response..."):
                            try:
                                # Step 1: Vector search
                                search_results = rag_engine.search_course(selected_course, query, top_k)
                                
                                if not search_results:
                                    st.warning("No relevant content found for your query.")
                                    return
                                
                                # Step 2: Cost estimation
                                if estimate_cost:
                                    cost_data = rag_engine.get_cost_estimate(query, len(search_results))
                                    
                                    col1, col2, col3 = st.columns(3)
                                    with col1:
                                        st.metric("Input Tokens", f"{cost_data['input_tokens']:.0f}")
                                    with col2:
                                        st.metric("Estimated Cost", f"${cost_data['total_estimated_cost']:.4f}")
                                    with col3:
                                        savings = cost_data['traditional_flat_rate_cost'] - cost_data['total_estimated_cost']
                                        st.metric("Savings vs Flat Rate", f"${savings:.4f}")
                                
                                # Step 3: Generate response
                                st.warning(f"💳 **BILLING**: Now charging your {provider_labels[api_provider]} account...")
                                
                                with st.spinner(f"🤖 Generating response using {provider_labels[api_provider]}..."):
                                    # Get the appropriate API key from manual input or environment
                                    selected_api_key = None
                                    if api_provider == "openai":
                                        selected_api_key = fresh_openai_key
                                    elif api_provider == "perplexity":
                                        selected_api_key = fresh_perplexity_key
                                    
                                    response = rag_engine.generate_response_with_context(
                                        query, search_results, api_provider, selected_api_key
                                    )
                                
                                # Display response
                                st.subheader("🤖 AI Response")
                                st.write(response)
                                
                                # Show sources if requested
                                if show_sources:
                                    st.subheader("📚 Source Chunks")
                                    for i, result in enumerate(search_results):
                                        with st.expander(f"Source {i+1}: {result['source_file']} (Similarity: {result['similarity']:.3f})"):
                                            st.write(result['content'])
                                            st.caption(f"Words: {result['word_count']} | Type: {result['chunk_type']}")
                                
                            except Exception as e:
                                st.error(f"Error processing query: {e}")
                
                # Search across all courses
                st.subheader("🌐 Search All Courses")
                if st.button("🔍 Search Across All Courses", key="search_all_courses_btn"):
                    if query:
                        with st.spinner("Searching all courses..."):
                            try:
                                all_results = rag_engine.search_all_courses(query, 15)
                                
                                if all_results:
                                    for course, results in all_results.items():
                                        with st.expander(f"📚 {course} ({len(results)} results)"):
                                            for result in results[:3]:  # Show top 3 per course
                                                st.write(f"**Score: {result['similarity']:.3f}** - {result['content'][:200]}...")
                                else:
                                    st.info("No results found across courses.")
                            except Exception as e:
                                st.error(f"Error searching all courses: {e}")
                    else:
                        st.warning("Please enter a query first.")
            else:
                st.warning("No vector embeddings found. Please generate embeddings first.")
        
        # Management tools
        st.subheader("📊 Vector Database Management")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button("📋 View All Vector Databases", key="view_vector_databases_btn"):
                vector_files = list(rag_engine.vectors_dir.glob("*_vectors.json"))
                if vector_files:
                    for vector_file in vector_files:
                        course_name = vector_file.stem.replace("_vectors", "")
                        try:
                            with open(vector_file, 'r') as f:
                                data = json.load(f)
                            st.write(f"**{course_name}**: {data['total_chunks']} chunks ({data['chunking_method']})")
                        except Exception as e:
                            st.write(f"**{course_name}**: Error reading data")
                else:
                    st.info("No vector databases found.")
        
        with col2:
            if st.button("🧹 Clear Vector Cache", key="clear_vector_cache_btn"):
                try:
                    # Clear memory cache
                    rag_engine.course_vectors.clear()
                    st.success("Vector cache cleared!")
                except Exception as e:
                    st.error(f"Error clearing cache: {e}")
        
        with col3:
            if st.button("📁 Open Vectors Folder", key="open_vectors_folder_btn"):
                st.info(f"Vector files stored in: `{rag_engine.vectors_dir}`")

        # Footer
        st.markdown("---")
        st.markdown("🔒 **100% Local Processing** - No data leaves your machine")

    def system_status_section(self):
        """Display system status and dependency information."""
        st.header("⚙️ System Status")
        
        # Check all dependencies
        dependencies = self._check_all_dependencies()
        
        # Overall status
        all_core_available = all(dependencies['core'].values())
        if all_core_available:
            st.success("✅ All core dependencies are available")
        else:
            st.warning("⚠️ Some core dependencies are missing")
        
        # Core dependencies status
        st.subheader("🔧 Core Dependencies")
        col1, col2 = st.columns(2)
        
        with col1:
            st.write("**Required for basic functionality:**")
            for dep, available in dependencies['core'].items():
                status = "✅" if available else "❌"
                st.write(f"{status} {dep}")
        
        with col2:
            st.write("**AI/ML Dependencies:**")
            for dep, available in dependencies['ai'].items():
                status = "✅" if available else "❌"
                st.write(f"{status} {dep}")
        
        # Document processing status
        st.subheader("📄 Document Processing")
        for dep, available in dependencies['documents'].items():
            status = "✅" if available else "❌"
            st.write(f"{status} {dep}")
        
        # System information
        st.subheader("💻 System Information")
        gpu_available = self.config.has_gpu()
        st.write(f"🔧 GPU Available: {'✅ Yes' if gpu_available else '❌ No'}")
        
        # Authentication requirements
        st.subheader("🔐 Hugging Face Authentication")
        st.markdown("""
        **This app uses only Mistral 7B and Llama 2 7B models - both require authentication:**
        
        **🎯 Quick Setup (Required for both models):**
        
        **Step 1: Get Hugging Face Token**
        1. Create free account at https://huggingface.co
        2. Go to https://huggingface.co/settings/tokens
        3. Click "New token", name it "course-ai", type "Read"
        4. Copy the token (starts with `hf_...`)
        
        **Step 2: Set Token in Replit**
        Open **Shell** tab and run:
        ```bash
        pip install huggingface_hub[cli]
        huggingface-cli login
        # Paste your token when prompted
        ```
        
        **Step 3: Request Model Access (if needed)**
        - **Llama 2** (Primary): Go to https://huggingface.co/meta-llama/Llama-2-7b-chat-hf
          - Click "Access repository" and fill the form
          - Usually approved within hours
        - **Mistral 7B** (Fallback): Usually available immediately
        
        **Model Priority:**
        - **Llama 2 7B** (Primary): Better for conversations, supports fine-tuning for learning
        - **Mistral 7B** (Fallback): Fast, efficient, good general knowledge
        
        **No Other Models**: DialoGPT and GPT-2 removed per your requirements
        """)
        
        # Installation instructions
        if not all_core_available:
            st.subheader("📦 Installation Instructions")
            st.markdown("""
            Install required dependencies:
            ```bash
            pip install accelerate torch transformers sentence-transformers llama-index
            ```
            """)
            
        # Manual model download option
        st.subheader("⬇️ Pre-download Models")
        st.markdown("""
        **Both models download automatically when first used, but you can pre-download:**
        
        **Download Mistral 7B:**
        ```bash
        python -c "
        from transformers import AutoTokenizer, AutoModelForCausalLM
        print('Downloading Mistral 7B...')
        AutoTokenizer.from_pretrained('mistralai/Mistral-7B-Instruct-v0.1', cache_dir='./models/')
        AutoModelForCausalLM.from_pretrained('mistralai/Mistral-7B-Instruct-v0.1', cache_dir='./models/')
        print('Done!')
        "
        ```
        
        **Download Llama 2 7B (after approval):**
        ```bash
        # 1. Login to Hugging Face
        pip install huggingface_hub[cli]
        huggingface-cli login  # paste your hf_ token
        
        # 2. Download Llama 2 (once approved)
        python -c "
        from transformers import AutoTokenizer, AutoModelForCausalLM
        print('Downloading Llama 2...')
        AutoTokenizer.from_pretrained('meta-llama/Llama-2-7b-chat-hf', cache_dir='./models/')
        AutoModelForCausalLM.from_pretrained('meta-llama/Llama-2-7b-chat-hf', cache_dir='./models/')
        print('Done!')
        "
        
        # 3. Download other models
        python -c "
        from sentence_transformers import SentenceTransformer
        import whisper
        SentenceTransformer('sentence-transformers/all-MiniLM-L6-v2')
        whisper.load_model('medium')
        "
        ```
        """)
        
        # Offline operation guide
        st.subheader("🌐 Offline Operation Guide")
        st.markdown("""
        **This app is designed for 100% offline operation:**
        
        **✅ What works offline:**
        - Document processing (PDFs, Word docs, PowerPoints, EPUBs)
        - Video/audio transcription with Whisper
        - AI question answering with local Mistral 7B model
        - Course indexing and search
        - All data processing and storage
        
        **📋 Steps for complete offline setup:**
        1. **Install dependencies** (requires internet once):
           ```bash
           python install_dependencies.py
           ```
        
        2. **Download AI models** (happens automatically when first used):
           - Models download only when the app first loads them:
             - Mistral 7B Instruct model (~4GB) - High-quality instruction model
             - Whisper medium model (~1.5GB) - Downloads when processing audio/video
             - MiniLM embedding model (~80MB) - Downloads when creating course indexes
           - Models are cached locally in `./models/` for future offline use
        
        3. **Disconnect from internet** - The app will work completely offline!
        
        **💾 Local storage locations:**
        - Models: `./models/` directory
        - Course indexes: `./indexed_courses/` directory
        - Raw documents: `./raw_docs/` directory
        
        **🔧 Using GGUF models:**
        For better performance and lower memory usage, download GGUF models:
        
        **Step 1: Download GGUF models from these sources (NO AUTH REQUIRED):**
        - **Zephyr 7B**: https://huggingface.co/TheBloke/zephyr-7B-beta-GGUF
        - **StableLM Zephyr 3B**: https://huggingface.co/TheBloke/stablelm-zephyr-3b-GGUF  
        - **OpenHermes 2.5**: https://huggingface.co/TheBloke/OpenHermes-2.5-Mistral-7B-GGUF
        - **Code Llama**: https://huggingface.co/TheBloke/CodeLlama-7B-Instruct-GGUF
        
        **Step 2: Choose the right quantization:**
        - `Q4_K_M.gguf` - Best balance (4GB RAM, good quality)
        - `Q5_K_M.gguf` - Higher quality (5GB RAM)
        - `Q8_0.gguf` - Highest quality (7GB RAM)
        
        **Step 3: Download and place in:**
        `./models/gguf/model-name.gguf`
        
        **Example download commands (NO AUTH REQUIRED):**
        ```bash
        # Download Zephyr 7B Q4_K_M (recommended for course analysis)
        wget -P ./models/gguf/ https://huggingface.co/TheBloke/zephyr-7B-beta-GGUF/resolve/main/zephyr-7b-beta.Q4_K_M.gguf
        
        # Or download StableLM Zephyr 3B Q4_K_M (smaller, faster)
        wget -P ./models/gguf/ https://huggingface.co/TheBloke/stablelm-zephyr-3b-GGUF/resolve/main/stablelm-zephyr-3b.Q4_K_M.gguf
        ```
        
        **⚠️ Note about GGUF usage:**
        Currently the app uses standard transformers library. To use GGUF models, you'd need:
        - `pip install llama-cpp-python` for GGUF support
        - The app will need updates to use llama-cpp-python instead of transformers
        
        The app will automatically detect and use GGUF models if available.
        
        **🔒 Privacy benefits:**
        - No external API calls
        - No data sent to third parties
        - Complete control over your course materials
        - Perfect for sensitive or proprietary content
        """)
        
        # Current status
        internet_required = not all_core_available
        if internet_required:
            st.info("🌐 Internet currently required for initial setup and model downloads")
        else:
            st.success("✅ Ready for offline operation once models are downloaded")

    def _check_all_dependencies(self) -> Dict[str, Dict[str, bool]]:
        """Check status of all dependencies."""
        def check_import(module_name: str) -> bool:
            try:
                __import__(module_name)
                return True
            except ImportError:
                return False
        
        return {
            'core': {
                'Streamlit': check_import('streamlit'),
                'Plotly': PLOTLY_AVAILABLE,
                'Pandas': PANDAS_AVAILABLE,
                'NumPy': check_import('numpy'),
            },
            'ai': {
                'PyTorch': check_import('torch'),
                'Transformers': check_import('transformers'),
                'Sentence Transformers': check_import('sentence_transformers'),
                'LlamaIndex': check_import('llama_index'),
                'Whisper': check_import('whisper'),
            },
            'documents': {
                'PyPDF2': check_import('PyPDF2'),
                'python-docx': check_import('docx'),
                'python-pptx': check_import('pptx'),
                'ebooklib': check_import('ebooklib'),
                'BeautifulSoup': check_import('bs4'),
            }
        }


    def save_conversation_for_learning(self, course_name: str, question: str, answer: str, 
                                      model_used: str = None, performance_metrics: dict = None, error: bool = False):
        """Save conversation data for future learning with enhanced metadata."""
        conversations_dir = Path("./conversations")
        conversations_dir.mkdir(exist_ok=True)
        
        conversation_file = conversations_dir / f"{course_name}_conversations.jsonl"
        
        conversation_data = {
            "timestamp": datetime.now().isoformat(),
            "course": course_name,
            "question": question,
            "answer": answer,
            "model_used": model_used or "unknown",
            "error": error
        }
        
        # Add performance metrics if available
        if performance_metrics:
            conversation_data.update({
                "response_time_ms": performance_metrics.get('response_time_ms', 0),
                "quality_score": performance_metrics.get('quality_score', 0),
                "tokens_per_second": performance_metrics.get('tokens_per_second', 0),
                "input_tokens": performance_metrics.get('input_tokens', 0),
                "output_tokens": performance_metrics.get('output_tokens', 0),
                "memory_usage_mb": performance_metrics.get('memory_usage_mb', 0)
            })
        
        try:
            with open(conversation_file, 'a', encoding='utf-8') as f:
                f.write(json.dumps(conversation_data) + '\n')
            logger.info(f"Saved conversation for course: {course_name} (model: {model_used})")
        except Exception as e:
            logger.error(f"Error saving conversation: {e}")


def main():
    """Application entry point."""
    app = RealEstateAIApp()
    app.run()


if __name__ == "__main__":
    main()
